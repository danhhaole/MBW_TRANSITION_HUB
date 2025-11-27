# Copyright (c) 2025, MBWCloud Co. and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
from frappe.utils import now, time_diff_in_seconds, flt
import json
from datetime import datetime, timedelta, date
from frappe import _
from frappe.utils.file_manager import save_file, get_file
from frappe.utils import now_datetime, now, getdate, today
from datetime import datetime
import re, io, os
import json
import pandas as pd
import time
from typing import Dict, List, Any, Optional, Set

class MiraImportsession(Document):
	def before_insert(self):
		"""Set default values before insert"""
		if not self.created_by:
			self.created_by = frappe.session.user
		if not self.started_at:
			self.started_at = now()
		if not self.status:
			self.status = "Pending"
		
		# Initialize logs storage
		self._temp_logs = []
	
	def before_save(self):
		"""Update calculated fields before save"""
		self.calculate_progress()
		self.estimate_time_remaining()
		
	def after_insert(self):
		"""Actions after document creation"""
		# Simple log without child table
		self.add_simple_log("info", f"Import session created for file: {self.file_name}")
		
	def calculate_progress(self):
		"""Calculate progress percentage"""
		if self.total_rows > 0:
			processed = (self.success_count or 0) + (self.failed_count or 0) + (self.duplicate_count or 0)
			self.progress_percentage = flt((processed / self.total_rows) * 100, 2)
		else:
			self.progress_percentage = 0
			
	def estimate_time_remaining(self):
		"""Estimate remaining time based on current progress"""
		if not self.started_at or self.status in ["Completed", "Failed", "Cancelled"]:
			self.estimated_time_remaining = ""
			return
			
		if self.progress_percentage <= 0:
			self.estimated_time_remaining = "Calculating..."
			return
			
		elapsed_time = time_diff_in_seconds(now(), self.started_at)
		if elapsed_time <= 0:
			self.estimated_time_remaining = "Calculating..."
			return
			
		# Calculate average time per row
		processed_rows = ((self.success_count or 0) + (self.failed_count or 0) + (self.duplicate_count or 0))
		if processed_rows <= 0:
			self.estimated_time_remaining = "Calculating..."
			return
			
		avg_time_per_row = elapsed_time / processed_rows
		remaining_rows = self.total_rows - processed_rows
		estimated_seconds = remaining_rows * avg_time_per_row
		
		if estimated_seconds <= 0:
			self.estimated_time_remaining = "Almost done"
		elif estimated_seconds < 60:
			self.estimated_time_remaining = f"{int(estimated_seconds)} seconds"
		elif estimated_seconds < 3600:
			minutes = int(estimated_seconds / 60)
			self.estimated_time_remaining = f"{minutes} minute(s)"
		else:
			hours = int(estimated_seconds / 3600)
			minutes = int((estimated_seconds % 3600) / 60)
			self.estimated_time_remaining = f"{hours}h {minutes}m"
	
	def add_simple_log(self, status, message, row_number=None, job_title=None,
					  document_name=None, processing_time=None):
		"""Add a simple log entry without child table"""
		try:
			# Initialize temp logs if not exists
			if not hasattr(self, '_temp_logs'):
				self._temp_logs = []
			
			log_entry = {
				"timestamp": now(),
				"status": status,
				"message": message,
				"row_number": row_number,
				"job_title": job_title,
				"document_name": document_name,
				"processing_time": processing_time
			}
			
			self._temp_logs.append(log_entry)
			
			# Store in error_summary as JSON (keep last 100 entries)
			if len(self._temp_logs) > 100:
				self._temp_logs = self._temp_logs[-100:]
			
			# Update error_summary field with logs
			try:
				existing_logs = []
				if self.error_summary:
					existing_logs = json.loads(self.error_summary)
				
				existing_logs.append(log_entry)
				if len(existing_logs) > 100:
					existing_logs = existing_logs[-100:]
				
				self.error_summary = json.dumps(existing_logs)
			except Exception as e:
				# If JSON parsing fails, start fresh
				self.error_summary = json.dumps([log_entry])
				
		except Exception as e:
			frappe.log_error(f"Failed to add log: {str(e)}", "ImportSession Log Error")
	
	def add_simple_log_candidate(self, status, message, row_number=None, full_name=None,
								  document_name=None, processing_time=None):
		"""Add a simple log entry without child table"""
		try:
			# Initialize temp logs if not exists
			if not hasattr(self, '_temp_logs_candidate'):
				self._temp_logs_candidate = []
			
			log_entry = {
				"timestamp": now(),
				"status": status,
				"message": message,
				"row_number": row_number,
				"full_name": full_name,
				"document_name": document_name,
				"processing_time": processing_time
			}
			
			self._temp_logs_candidate.append(log_entry)
			
			# Store in error_summary as JSON (keep last 100 entries)
			if len(self._temp_logs_candidate) > 100:
				self._temp_logs_candidate = self._temp_logs_candidate[-100:]
			
			# Update error_summary field with logs
			try:
				existing_logs = []
				if self.error_summary_candidate:
					existing_logs = json.loads(self.error_summary_candidate)
				
				existing_logs.append(log_entry)
				if len(existing_logs) > 100:
					existing_logs = existing_logs[-100:]
				
				self.error_summary_candidate = json.dumps(existing_logs)
			except Exception as e:
				# If JSON parsing fails, start fresh
				self.error_summary_candidate = json.dumps([log_entry])
				
		except Exception as e:
			frappe.log_error(f"Failed to add log: {str(e)}", "ImportSession Log Error")
	
	def add_simple_log_talent(self, status, message, row_number=None, full_name=None,
								  document_name=None, processing_time=None):
		"""Add a simple log entry without child table"""
		try:
			# Initialize temp logs if not exists
			if not hasattr(self, '_temp_logs_talent'):
				self._temp_logs_talent = []
			
			log_entry = {
				"timestamp": now(),
				"status": status,
				"message": message,
				"row_number": row_number,
				"full_name": full_name,
				"document_name": document_name,
				"processing_time": processing_time
			}
			
			self._temp_logs_talent.append(log_entry)
			
			# Store in error_summary as JSON (keep last 100 entries)
			if len(self._temp_logs_talent) > 100:
				self._temp_logs_talent = self._temp_logs_talent[-100:]
			
			# Update error_summary field with logs
			try:
				existing_logs = []
				if self.error_summary:
					existing_logs = json.loads(self.error_summary)
				
				existing_logs.append(log_entry)
				if len(existing_logs) > 100:
					existing_logs = existing_logs[-100:]
				
				self.error_summary = json.dumps(existing_logs)
			except Exception as e:
				# If JSON parsing fails, start fresh
				self.error_summary = json.dumps([log_entry])
				
		except Exception as e:
			frappe.log_error(f"Failed to add log: {str(e)}", "ImportSession Log Error")

	def update_status(self, status, save=True):
		"""Update session status"""
		old_status = self.status
		self.status = status
		
		if status == "Completed":
			self.completed_at = now()
			self.progress_percentage = 100
			
		if save:
			try:
				self.save(ignore_permissions=True)
				frappe.db.commit()
			except Exception as e:
				frappe.log_error(f"Failed to save session: {str(e)}", "Session Save Error")
			
		# Publish real-time update
		try:
			frappe.publish_realtime(
				'import_session_update',
				{
					'session_id': self.name,
					'status': status,
					'old_status': old_status,
					'progress': self.progress_percentage,
					'success_count': self.success_count,
					'failed_count': self.failed_count,
					'duplicate_count': self.duplicate_count
				}
			)
		except Exception as e:
			frappe.log_error(f"Failed to publish update: {str(e)}", "Realtime Update Error")
		
	def mark_success(self, row_number=None, job_title=None, document_name=None, 
					processing_time=None):
		"""Mark a row as successfully processed"""
		self.success_count = (self.success_count or 0) + 1
		self.add_simple_log(
			status="success",
			message=f"Successfully created job opening: {document_name}",
			row_number=row_number,
			job_title=job_title,
			document_name=document_name,
			processing_time=processing_time
		)
	
	def mark_success_candidate(self, row_number=None, full_name=None, document_name=None, 
					processing_time=None):
		"""Mark a row as successfully processed"""
		self.success_count = (self.success_count or 0) + 1
		self.add_simple_log_candidate(
			status="success",
			message=f"Successfully created candidate: {document_name}",
			row_number=row_number,
			full_name=full_name,
			document_name=document_name,
			processing_time=processing_time
		)
	
	def mark_success_talent(self, row_number=None, full_name=None, document_name=None, 
					processing_time=None):
		"""Mark a row as successfully processed"""
		self.success_count = (self.success_count or 0) + 1
		self.add_simple_log_talent(
			status="success",
			message=f"Successfully created talent: {document_name}",
			row_number=row_number,
			full_name=full_name,
			document_name=document_name,
			processing_time=processing_time
		)
		
	def mark_failure(self, row_number=None, job_title=None, error_message=None, 
					processing_time=None):
		"""Mark a row as failed"""
		self.failed_count = (self.failed_count or 0) + 1
		self.add_simple_log(
			status="error",
			message=error_message or "Processing failed",
			row_number=row_number,
			job_title=job_title,
			processing_time=processing_time
		)

	def mark_failure_candidate(self, row_number=None, full_name=None, error_message=None, 
					processing_time=None):
		"""Mark a row as failed"""
		self.failed_count = (self.failed_count or 0) + 1
		self.add_simple_log_candidate(
			status="error",
			message=error_message or "Processing failed",
			row_number=row_number,
			full_name=full_name,
			processing_time=processing_time
		)
	
	def mark_failure_talent(self, row_number=None, full_name=None, error_message=None, 
					processing_time=None):
		"""Mark a row as failed"""
		self.failed_count = (self.failed_count or 0) + 1
		self.add_simple_log_talent(
			status="error",
			message=error_message or "Processing failed",
			row_number=row_number,
			full_name=full_name,
			processing_time=processing_time
		)
		
	def mark_duplicate(self, row_number=None, job_title=None, message=None, 
					  processing_time=None):
		"""Mark a row as duplicate"""
		self.duplicate_count = (self.duplicate_count or 0) + 1
		self.add_simple_log(
			status="skipped",
			message=message or "Duplicate job opening skipped",
			row_number=row_number,
			job_title=job_title,
			processing_time=processing_time
		)
		
	def get_summary(self):
		"""Get import summary"""
		return {
			"session_id": self.name,
			"file_name": self.file_name,
			"status": self.status,
			"total_rows": self.total_rows,
			"success_count": self.success_count or 0,
			"failed_count": self.failed_count or 0,
			"duplicate_count": self.duplicate_count or 0,
			"progress_percentage": self.progress_percentage or 0,
			"started_at": self.started_at,
			"completed_at": self.completed_at,
			"estimated_time_remaining": self.estimated_time_remaining,
			"field_mapping": json.loads(self.field_mapping) if self.field_mapping else {}
		}
		
	def get_logs(self, limit=50):
		"""Get import logs"""
		logs = []
		if self.error_summary:
			try:
				all_logs = json.loads(self.error_summary)
				logs = all_logs[-limit:] if all_logs else []
			except:
				logs = []
		return logs
		
	def cancel_import(self, reason=None):
		"""Cancel the import session"""
		self.status = "Cancelled"
		self.completed_at = now()
		self.add_simple_log(
			status="info", 
			message=f"Import cancelled. Reason: {reason or 'User cancelled'}"
		)
		self.save(ignore_permissions=True)
		
		# Publish cancellation
		frappe.publish_realtime(
			'import_cancelled',
			{
				'session_id': self.name,
				'reason': reason
			}
		)

# Candidate
@frappe.whitelist()
@frappe.whitelist()
def download_candidate_template():
	"""Generate and download import template Excel file for Candidate"""
	try:
		# Get Mira Candidate fields
		meta = frappe.get_meta("Mira Candidate")
		
		# Define important fields in order
		important_fields = [
			'full_name', 'email', 'phone', 'headline', 'source',
			'cv_original_url', 'skills', 'status'
		]
		
		template_fields = []
		# Add important fields first
		for fieldname in important_fields:
			field = meta.get_field(fieldname)
			if field:
				template_fields.append({
					'fieldname': field.fieldname,
					'label': field.label,
					'fieldtype': field.fieldtype,
					'required': field.reqd if hasattr(field, 'reqd') else False,
					'options': field.options if hasattr(field, 'options') else ''
				})
		
		# Create template DataFrame with headers and sample data
		headers = []
		sample_data = {}
		
		for field in template_fields:
			header = f"{field['label']}"
			if field.get('required'):
				header += " *"
			headers.append(header)
			
			# Add sample data
			if field['fieldname'] == 'full_name':
				sample_data[header] = 'Nguyễn Văn A'
			elif field['fieldname'] == 'email':
				sample_data[header] = 'nguyen.van.a@example.com'
			elif field['fieldname'] == 'phone':
				sample_data[header] = '091234567823'
			elif field['fieldname'] == 'headline':
				sample_data[header] = 'Senior Software Engineer at ABC Company'
			elif field['fieldname'] == 'source':
				sample_data[header] = 'LinkedIn_Sourcing'
			elif field['fieldname'] == 'cv_original_url':
				sample_data[header] = 'https://example.com/cv/nguyen-van-a.pdf'
			elif field['fieldname'] == 'skills':
				sample_data[header] = 'Python, JavaScript, React, Node.js'
			elif field['fieldname'] == 'status':
				sample_data[header] = 'NEW'
			else:
				sample_data[header] = ''
		
		# Create DataFrame with sample data
		df = pd.DataFrame([sample_data])
		
		# Save to temporary file and return
		import tempfile
		
		with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
			with pd.ExcelWriter(tmp.name, engine='openpyxl') as writer:
				df.to_excel(writer, sheet_name='Candidate Template', index=False)
				
				# Get workbook and worksheet for styling
				workbook = writer.book
				worksheet = writer.sheets['Candidate Template']
				
				# Style headers
				from openpyxl.styles import Font, PatternFill, Alignment
				from openpyxl.utils import get_column_letter
				
				header_font = Font(bold=True, color="FFFFFF")
				header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
				center_alignment = Alignment(horizontal="center", vertical="center")
				
				# Apply styling to header row
				for col_num, header in enumerate(headers, 1):
					cell = worksheet.cell(row=1, column=col_num)
					cell.font = header_font
					cell.fill = header_fill
					cell.alignment = center_alignment
				
				# 👉 Giữ nguyên số 0 đầu của số điện thoại bằng cách định dạng cột "Phone" là text
				for col_num, header in enumerate(headers, 1):
					column_letter = get_column_letter(col_num)
					
					# Nếu header có chữ "phone" => format text
					if 'phone' in header.lower():
						for row in range(2, len(df) + 2):  # từ hàng 2 (dữ liệu)
							cell = worksheet.cell(row=row, column=col_num)
							cell.number_format = '@'  # '@' = định dạng text
					
					# Thiết lập độ rộng cột
					worksheet.column_dimensions[column_letter].width = max(len(str(header)) + 5, 15)
			
			# Read file content
			with open(tmp.name, 'rb') as f:
				content = f.read()
			
			# Clean up
			os.unlink(tmp.name)
		
		# Create file document
		file_doc = frappe.get_doc({
			"doctype": "File",
			"file_name": f"candidate_import_template_{frappe.utils.now_datetime().strftime('%Y%m%d_%H%M%S')}.xlsx",
			"content": content,
			"is_private": 0
		})
		file_doc.save(ignore_permissions=True)
		
		return {
			"file_url": file_doc.file_url,
			"file_name": file_doc.file_name
		}
		
	except Exception as e:
		frappe.log_error(frappe.get_traceback(), "Candidate Template Generation Error")
		frappe.throw(_("Error generating candidate template: {0}").format(str(e)))

@frappe.whitelist()
def upload_and_preview_candidate():
	"""Enhanced preview endpoint with better error handling and encoding support for Candidate"""
	try:
		uploaded = frappe.request.files.get('file')
		if not uploaded:
			frappe.throw(_("No file uploaded"))

		filename = uploaded.filename or ""
		ext = filename.split(".")[-1].lower()

		# Validate file type
		if ext not in ['csv', 'xls', 'xlsx']:
			frappe.throw(_("Unsupported file type. Only CSV, XLS, XLSX files are allowed."))

		try:
			file_bytes = uploaded.read()
			bio = io.BytesIO(file_bytes)
			
			df = None
			if ext == "csv":
				# Try multiple encodings for CSV
				encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
				for encoding in encodings:
					try:
						bio.seek(0)
						# 👉 Force all columns to string
						df = pd.read_csv(bio, encoding=encoding, dtype=str, keep_default_na=False)
						break
					except (UnicodeDecodeError, UnicodeError):
						continue
				
				if df is None:
					frappe.throw(_("Unable to read CSV file. Please check file encoding."))
					
			elif ext in ("xls", "xlsx"):
				try:
					# 👉 Force all cells to be read as strings to keep leading zeros (e.g. phone numbers)
					df = pd.read_excel(bio, dtype=str, engine='openpyxl' if ext == 'xlsx' else None)
				except Exception as e:
					frappe.throw(_("Error reading Excel file: {0}").format(str(e)))

			# Validate DataFrame
			if df is None or df.empty:
				frappe.throw(_("File is empty or cannot be read"))

			# Clean column names and remove empty rows
			df.columns = [str(col).strip() for col in df.columns]
			df = df.dropna(how='all')
			
			if df.empty:
				frappe.throw(_("No valid data found in file"))

			# ✅ Convert all values to string (avoid float/int converting again)
			df = df.applymap(lambda x: str(x).strip() if pd.notna(x) else "")

		except Exception as e:
			frappe.log_error(frappe.get_traceback(), "Candidate Import Preview Read Error")
			frappe.throw(_("Failed to read file: {0}").format(str(e)))

		# Prepare response data
		columns = [str(c) for c in df.columns.tolist()]
		sample = df.head(10).fillna("").to_dict(orient="records")

		# Get available fields from Mira Candidate for mapping suggestions
		meta = frappe.get_meta("Mira Candidate")
		available_fields = []
		for field in meta.fields:
			if field.fieldtype not in ['Section Break', 'Column Break', 'Tab Break', 'HTML']:
				available_fields.append({
					'fieldname': field.fieldname,
					'label': field.label,
					'fieldtype': field.fieldtype,
					'reqd': field.reqd
				})

		return {
			"filename": filename,
			"columns": columns,
			"sample": sample,
			"total_rows": len(df),
			"available_fields": available_fields
		}
		
	except Exception as e:
		frappe.log_error(frappe.get_traceback(), "Candidate Upload Preview Error")
		frappe.throw(_("Preview failed: {0}").format(str(e)))

@frappe.whitelist()
def import_with_mapping_candidate():
	"""Enhanced import with mapping using Mira Importsession for tracking (Candidate)"""
	try:
		# 1️⃣ Nhận dữ liệu từ form
		uploaded = frappe.request.files.get('file')
		mapping_raw = frappe.form_dict.get('mapping', '{}')
		selected_job_opening = frappe.form_dict.get('selected_job_opening') or frappe.form_dict.get('job_opening')
		batch_size = int(frappe.form_dict.get('batch_size', 100))
		validate_only = frappe.form_dict.get('validate_only', '0') == '1'
		skip_duplicates = frappe.form_dict.get('skip_duplicates', '1') == '1'

		if not uploaded:
			frappe.throw(_("No file uploaded"))

		# 2️⃣ Lưu file tạm vào Frappe File
		file_doc = save_file(
			fname=uploaded.filename,
			content=uploaded.read(),
			dt="Mira Importsession",
			dn=uploaded.filename,
			is_private=0
		)
		try:
			# 3️⃣ Đọc file vào dataframe
			df = read_file_to_dataframe(file_doc.get_full_path(), uploaded.filename)

			# 4️⃣ Parse mapping
			mapping = json.loads(mapping_raw)
			if not mapping:
				frappe.throw(_("Field mapping is required"))

			# 5️⃣ Chuẩn hoá mapping (Excel header → field hệ thống)
			normalized_mapping = {}
			for source_col, target_field in mapping.items():
				if target_field:
					normalized_mapping[str(source_col).strip().lower()] = str(target_field).strip()

			# 6️⃣ Tạo import session
			session = create_import_session_candidate(
				file_name=uploaded.filename,
				file_url=file_doc.file_url,
				total_rows=len(df),
				import_type="Candidate",
				selected_job_opening=selected_job_opening,
				field_mapping=normalized_mapping,  # mapping chiều đúng
				batch_size=batch_size,
				validate_only=validate_only,
				skip_duplicates=skip_duplicates
			)
			session.add_simple_log_candidate("info", f"Import session started for {len(df)} rows")
			session.update_status("Processing")

			# 7️⃣ Chạy chế độ validate hoặc insert
			if validate_only:
				results = validate_import_data_with_session_candidate(df, normalized_mapping, session)
			else:
				results = process_import_data_with_session_candidate(df, normalized_mapping, selected_job_opening, session, batch_size)

			# 9️⃣ Cập nhật trạng thái session
			if results.get("failed", 0) == 0:
				session.update_status("Completed")
			else:
				session.update_status("Completed")
				session.error_summary = f"{results['failed']} rows failed during import"

			return {
				**results,
				"session_id": session.name
			}

		finally:
			# 🔟 Xoá file tạm
			try:
				frappe.delete_doc("File", file_doc.name, ignore_permissions=True)
			except Exception as cleanup_err:
				print(f"Failed to delete temp file: {str(cleanup_err)}")

	except Exception as e:
		frappe.log_error(frappe.get_traceback(), "Talent Import Error")
		frappe.throw(_("Import failed: {0}").format(str(e)))



# Job
@frappe.whitelist()
def download_import_template():
	"""Generate and download import template Excel file for Job Opening"""
	try:
		# Get JobOpening fields
		meta = frappe.get_meta("Mira Job Opening")
		
		# Define important fields in order (excluding job_code as it's auto-generated)
		important_fields = [
			'job_title', 'description', 'requirements', 'benefits',
			'department_name', 'location_name', 'number_of_openings', 
			'posting_date', 'closing_date', 'approval_status', 'owner_id'
		]
		
		template_fields = []
		# Add important fields first
		for fieldname in important_fields:
			field = meta.get_field(fieldname)
			if field:
				template_fields.append({
					'fieldname': field.fieldname,
					'label': field.label,
					'fieldtype': field.fieldtype,
					'required': field.reqd,
					'options': field.options
				})
		
		# Create template DataFrame with headers and sample data
		headers = []
		sample_data = {}
		
		for field in template_fields:
			header = f"{field['label']}"
			if field['required']:
				header += " *"
			headers.append(header)
			
			# Add sample data
			if field['fieldname'] == 'job_title':
				sample_data[header] = 'Software Engineer'
			elif field['fieldname'] == 'description':
				sample_data[header] = 'Develop and maintain software applications'
			elif field['fieldname'] == 'requirements':
				sample_data[header] = 'Bachelor degree in Computer Science, 2+ years experience'
			elif field['fieldname'] == 'benefits':
				sample_data[header] = 'Competitive salary, health insurance, flexible hours'
			elif field['fieldname'] == 'department_name':
				sample_data[header] = 'Engineering'
			elif field['fieldname'] == 'location_name':
				sample_data[header] = 'Ho Chi Minh City'
			elif field['fieldname'] == 'number_of_openings':
				sample_data[header] = '2'
			elif field['fieldname'] == 'posting_date':
				sample_data[header] = '2024-01-01'
			elif field['fieldname'] == 'closing_date':
				sample_data[header] = '2024-02-01'
			elif field['fieldname'] == 'approval_status':
				sample_data[header] = 'Draft'
			elif field['fieldname'] == 'owner_id':
				sample_data[header] = 'Administrator'
			else:
				sample_data[header] = ''
		
		# Create DataFrame with sample data
		df = pd.DataFrame([sample_data])
		
		# Save to temporary file and return
		import tempfile
		
		with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
			with pd.ExcelWriter(tmp.name, engine='openpyxl') as writer:
				df.to_excel(writer, sheet_name='Job Opening Template', index=False)
				
				# Get workbook and worksheet for styling
				workbook = writer.book
				worksheet = writer.sheets['Job Opening Template']
				
				# Style headers
				from openpyxl.styles import Font, PatternFill, Alignment
				
				header_font = Font(bold=True, color="FFFFFF")
				header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
				center_alignment = Alignment(horizontal="center", vertical="center")
				
				# Apply styling to header row
				for col_num, header in enumerate(headers, 1):
					cell = worksheet.cell(row=1, column=col_num)
					cell.font = header_font
					cell.fill = header_fill
					cell.alignment = center_alignment
				
				# Set column widths
				for col_num, header in enumerate(headers, 1):
					column_letter = worksheet.cell(row=1, column=col_num).column_letter
					worksheet.column_dimensions[column_letter].width = max(len(header) + 5, 15)
			
			# Read file content
			with open(tmp.name, 'rb') as f:
				content = f.read()
			
			# Clean up
			os.unlink(tmp.name)
		
		# Create file document
		file_doc = frappe.get_doc({
			"doctype": "File",
			"file_name": f"job_opening_import_template_{frappe.utils.now_datetime().strftime('%Y%m%d_%H%M%S')}.xlsx",
			"content": content,
			"is_private": 0
		})
		file_doc.save(ignore_permissions=True)
		
		return {
			"file_url": file_doc.file_url,
			"file_name": file_doc.file_name
		}
		
	except Exception as e:
		frappe.log_error(frappe.get_traceback(), "Job Opening Template Generation Error")
		frappe.throw(_("Failed to generate template: {0}").format(str(e)))

@frappe.whitelist()
def upload_and_preview():
	"""Enhanced preview endpoint with better error handling and encoding support for Job Opening"""
	try:
		uploaded = frappe.request.files.get('file')
		if not uploaded:
			frappe.throw(_("No file uploaded"))

		filename = uploaded.filename or ""
		ext = filename.split(".")[-1].lower()

		# Validate file type
		if ext not in ['csv', 'xls', 'xlsx']:
			frappe.throw(_("Unsupported file type. Only CSV, XLS, XLSX files are allowed."))

		# Read file into DataFrame with improved error handling
		try:
			file_bytes = uploaded.read()
			bio = io.BytesIO(file_bytes)
			
			df = None
			if ext == "csv":
				# Try multiple encodings for CSV
				encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
				for encoding in encodings:
					try:
						bio.seek(0)
						df = pd.read_csv(bio, encoding=encoding)
						break
					except (UnicodeDecodeError, UnicodeError):
						continue
				
				if df is None:
					frappe.throw(_("Unable to read CSV file. Please check file encoding."))
					
			elif ext in ("xls", "xlsx"):
				try:
					df = pd.read_excel(bio, engine='openpyxl' if ext == 'xlsx' else None)
				except Exception as e:
					frappe.throw(_("Error reading Excel file: {0}").format(str(e)))

			# Validate DataFrame
			if df is None or df.empty:
				frappe.throw(_("File is empty or cannot be read"))

			# Clean column names and remove empty rows
			df.columns = [str(col).strip() for col in df.columns]
			df = df.dropna(how='all')
			
			if df.empty:
				frappe.throw(_("No valid data found in file"))

		except Exception as e:
			frappe.log_error(frappe.get_traceback(), "Job Opening Import Preview Read Error")
			frappe.throw(_("Failed to read file: {0}").format(str(e)))

		# Prepare response data
		columns = [str(c) for c in df.columns.tolist()]
		sample = df.head(10).fillna("").to_dict(orient="records")

		# Get available fields from JobOpening for mapping suggestions
		meta = frappe.get_meta("Mira Job Opening")
		available_fields = []
		for field in meta.fields:
			if field.fieldtype not in ['Section Break', 'Column Break', 'Tab Break', 'HTML']:
				available_fields.append({
					'fieldname': field.fieldname,
					'label': field.label,
					'fieldtype': field.fieldtype,
					'reqd': field.reqd
				})

		return {
			"filename": filename,
			"columns": columns,
			"sample": sample,
			"total_rows": len(df),
			"available_fields": available_fields
		}
		
	except Exception as e:
		frappe.log_error(frappe.get_traceback(), "Job Opening Upload Preview Error")
		frappe.throw(_("Preview failed: {0}").format(str(e)))

@frappe.whitelist()
def import_with_mapping():
	"""Enhanced import with mapping using Mira Importsession for tracking (Job Opening)"""
	try:
		# Get form data
		uploaded = frappe.request.files.get('file')
		mapping_raw = frappe.form_dict.get('mapping', '{}')
		selected_job_opening = frappe.form_dict.get('selected_job_opening') or frappe.form_dict.get('job_opening')
		batch_size = int(frappe.form_dict.get('batch_size', 100))
		validate_only = frappe.form_dict.get('validate_only', '0') == '1'
		skip_duplicates = frappe.form_dict.get('skip_duplicates', '1') == '1'

		if not uploaded:
			frappe.throw(_("No file uploaded"))

		# Save file to Frappe
		file_doc = save_file(
			fname=uploaded.filename,
			content=uploaded.read(),
			dt="Mira Importsession",
			dn=uploaded.filename,
			is_private=0
		)

		try:
			# Read file with pandas
			df = read_file_to_dataframe(file_doc.get_full_path(), uploaded.filename)
			# Process mapping
			mapping = json.loads(mapping_raw)
			if not mapping:
				frappe.throw(_("Field mapping is required"))

			# Create import session
			session = create_import_session(
				file_name=uploaded.filename,
				file_url=file_doc.file_url,
				total_rows=len(df),
				import_type="Job Opening",
				selected_job_opening=selected_job_opening,
				field_mapping=mapping,
				batch_size=batch_size,
				validate_only=validate_only,
				skip_duplicates=skip_duplicates
			)
			session.add_simple_log("info", f"Import session started for {len(df)} rows")
			session.update_status("Processing")

			if validate_only:
				# Only validate, don't insert
				results = validate_import_data_with_session(df, mapping, selected_job_opening, session)
			else:
				# Process and insert data
				results = process_import_data_with_session(df, mapping, selected_job_opening, session, batch_size)

			# Update session with final results
			if results.get("failed", 0) == 0:
				session.update_status("Completed")
			else:
				session.update_status("Completed")
				session.error_summary = f"{results['failed']} rows failed during import"

			return {
				**results,
				"session_id": session.name
			}

		finally:
			# Clean up temporary file
			try:
				frappe.delete_doc("File", file_doc.name, ignore_permissions=True)
			except:
				pass

	except Exception as e:
		frappe.log_error(frappe.get_traceback(), "Job Opening Import Error")
		frappe.throw(_("Import failed: {0}").format(str(e)))

# Talent
@frappe.whitelist()
def download_talent_template():
	"""Generate and download import template Excel file for Talent"""
	try:
		# Get MiraTalent fields
		meta = frappe.get_meta("Mira Talent")
		
		# Define important fields in order
		important_fields = [
			'full_name', 'email', 'phone', 'skills', 'source', 'crm_status',
			'availability_date', 'expected_salary', 'hard_skills', 'soft_skills',
			'domain_expertise', 'cultural_fit', 'internal_rating', 'recruitment_readiness'
		]
		
		template_fields = []
		# Add important fields
		for fieldname in important_fields:
			field = meta.get_field(fieldname)
			if field:
				template_fields.append({
					'fieldname': field.fieldname,
					'label': field.label,
					'fieldtype': field.fieldtype,
					'required': field.reqd or fieldname in ['full_name', 'email'],
					'options': field.options
				})
		
		# Create template DataFrame with headers and sample data
		headers = []
		sample_data = {}
		
		for field in template_fields:
			header = f"{field['label']}"
			if field['required']:
				header += " *"
			headers.append(header)
			
			# Add sample data
			if field['fieldname'] == 'full_name':
				sample_data[header] = 'Nguyen Van A'
			elif field['fieldname'] == 'email':
				sample_data[header] = 'nguyenvana@example.com'
			elif field['fieldname'] == 'phone':
				sample_data[header] = '091234567823'
			elif field['fieldname'] == 'skills':
				sample_data[header] = 'Python, JavaScript, SQL'
			elif field['fieldname'] == 'source':
				sample_data[header] = 'Import Excel'
			elif field['fieldname'] == 'crm_status':
				sample_data[header] = 'New'
			elif field['fieldname'] == 'availability_date':
				sample_data[header] = '2024-03-01'
			elif field['fieldname'] == 'expected_salary':
				sample_data[header] = '25000000'
			elif field['fieldname'] == 'hard_skills':
				sample_data[header] = 'Python, React, Node.js, PostgreSQL'
			elif field['fieldname'] == 'soft_skills':
				sample_data[header] = 'Communication, Leadership, Problem Solving'
			elif field['fieldname'] == 'domain_expertise':
				sample_data[header] = 'Fintech, E-commerce, Web Development'
			elif field['fieldname'] == 'cultural_fit':
				sample_data[header] = 'High'
			elif field['fieldname'] == 'internal_rating':
				sample_data[header] = 'A'
			elif field['fieldname'] == 'recruitment_readiness':
				sample_data[header] = 'Warm'
			else:
				sample_data[header] = ''
		
		# Create DataFrame with sample data
		df = pd.DataFrame([sample_data])
		
		# Save to temporary file and return
		import tempfile
		import os
		
		with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
			with pd.ExcelWriter(tmp.name, engine='openpyxl') as writer:
				df.to_excel(writer, sheet_name='Talent Import Template', index=False)
				
				# Get workbook and worksheet for styling
				workbook = writer.book
				worksheet = writer.sheets['Talent Import Template']
				
				# Style headers
				from openpyxl.styles import Font, PatternFill, Alignment
				
				header_font = Font(bold=True, color="FFFFFF")
				header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
				center_alignment = Alignment(horizontal="center", vertical="center")
				
				# Apply styling to header row
				for col_num, header in enumerate(headers, 1):
					cell = worksheet.cell(row=1, column=col_num)
					cell.font = header_font
					cell.fill = header_fill
					cell.alignment = center_alignment
				
				# Set column widths
				for col_num, header in enumerate(headers, 1):
					column_letter = worksheet.cell(row=1, column=col_num).column_letter
					worksheet.column_dimensions[column_letter].width = max(len(str(header)) + 5, 15)
			
			# Read file content
			with open(tmp.name, 'rb') as f:
				content = f.read()
			
			# Clean up
			os.unlink(tmp.name)
		
		# Create file document
		file_doc = frappe.get_doc({
			"doctype": "File",
			"file_name": f"talent_import_template_{frappe.utils.now_datetime().strftime('%Y%m%d_%H%M%S')}.xlsx",
			"content": content,
			"is_private": 0
		})
		file_doc.save(ignore_permissions=True)
		
		return {
			"file_url": file_doc.file_url,
			"file_name": file_doc.file_name
		}
		
	except Exception as e:
		frappe.log_error(frappe.get_traceback(), "Talent Template Generation Error")
		frappe.throw(_("Failed to generate template: {0}").format(str(e)))

@frappe.whitelist()
def upload_and_preview_talent():
	"""Enhanced preview endpoint with better error handling and encoding support for Candidate"""
	try:
		uploaded = frappe.request.files.get('file')
		if not uploaded:
			frappe.throw(_("No file uploaded"))

		filename = uploaded.filename or ""
		ext = filename.split(".")[-1].lower()

		# Validate file type
		if ext not in ['csv', 'xls', 'xlsx']:
			frappe.throw(_("Unsupported file type. Only CSV, XLS, XLSX files are allowed."))

		try:
			file_bytes = uploaded.read()
			bio = io.BytesIO(file_bytes)
			
			df = None
			if ext == "csv":
				# Try multiple encodings for CSV
				encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
				for encoding in encodings:
					try:
						bio.seek(0)
						# 👉 Force all columns to string
						df = pd.read_csv(bio, encoding=encoding, dtype=str, keep_default_na=False)
						break
					except (UnicodeDecodeError, UnicodeError):
						continue
				
				if df is None:
					frappe.throw(_("Unable to read CSV file. Please check file encoding."))
					
			elif ext in ("xls", "xlsx"):
				try:
					# 👉 Force all cells to be read as strings to keep leading zeros (e.g. phone numbers)
					df = pd.read_excel(bio, dtype=str, engine='openpyxl' if ext == 'xlsx' else None)
				except Exception as e:
					frappe.throw(_("Error reading Excel file: {0}").format(str(e)))

			# Validate DataFrame
			if df is None or df.empty:
				frappe.throw(_("File is empty or cannot be read"))

			# Clean column names and remove empty rows
			df.columns = [str(col).strip() for col in df.columns]
			df = df.dropna(how='all')
			
			if df.empty:
				frappe.throw(_("No valid data found in file"))

			# ✅ Convert all values to string (avoid float/int converting again)
			df = df.applymap(lambda x: str(x).strip() if pd.notna(x) else "")

		except Exception as e:
			frappe.log_error(frappe.get_traceback(), "Talent Import Preview Read Error")
			frappe.throw(_("Failed to read file: {0}").format(str(e)))

		# Prepare response data
		columns = [str(c) for c in df.columns.tolist()]
		sample = df.head(10).fillna("").to_dict(orient="records")

		# Get available fields from Mira Candidate for mapping suggestions
		meta = frappe.get_meta("Mira Talent")
		available_fields = []
		for field in meta.fields:
			if field.fieldtype not in ['Section Break', 'Column Break', 'Tab Break', 'HTML']:
				available_fields.append({
					'fieldname': field.fieldname,
					'label': field.label,
					'fieldtype': field.fieldtype,
					'reqd': field.reqd
				})

		return {
			"filename": filename,
			"columns": columns,
			"sample": sample,
			"total_rows": len(df),
			"available_fields": available_fields
		}
		
	except Exception as e:
		frappe.log_error(frappe.get_traceback(), "Talent Upload Preview Error")
		frappe.throw(_("Preview failed: {0}").format(str(e)))

@frappe.whitelist()
def import_with_mapping_talent():
	"""Enhanced import with mapping using Mira Importsession for tracking"""
	try:
		print(">>>>>>>>>>>>>>>>>>> 1.")
		# 1️⃣ Nhận dữ liệu từ form
		uploaded = frappe.request.files.get('file')
		mapping_raw = frappe.form_dict.get('mapping', '{}')
		batch_size = int(frappe.form_dict.get('batch_size', 100))
		validate_only = frappe.form_dict.get('validate_only', '0') == '1'
		skip_duplicates = frappe.form_dict.get('skip_duplicates', '1') == '1'
		segment_id = frappe.form_dict.get('segment_id')  # Get segment_id if provided

		if not uploaded:
			frappe.throw(_("No file uploaded"))

		print(">>>>>>>>>>>>>>>>>>> 2.")
		# 2️⃣ Lưu file tạm vào Frappe File
		file_doc = save_file(
			fname=uploaded.filename,
			content=uploaded.read(),
			dt="Mira Importsession",
			dn=uploaded.filename,
			is_private=0
		)

		try:
			print(">>>>>>>>>>>>>>>>>>> 3.")
			# 3️⃣ Đọc file vào dataframe
			df = read_file_to_dataframe(file_doc.get_full_path(), uploaded.filename)

			print(">>>>>>>>>>>>>>>>>>> 4.")
			# 4️⃣ Parse mapping
			mapping = json.loads(mapping_raw)
			if not mapping:
				frappe.throw(_("Field mapping is required"))

			print(">>>>>>>>>>>>>>>>>>> 5.")
			# 5️⃣ Chuẩn hoá mapping (Excel header → field hệ thống)
			normalized_mapping = {}
			for source_col, target_field in mapping.items():
				if target_field:
					normalized_mapping[str(source_col).strip().lower()] = str(target_field).strip()

			print(">>>>>>>>>>>>>>>>>>> 6.")
			# 6️⃣ Tạo import session
			session = create_import_session_talennt(
				file_name=uploaded.filename,
				file_url=file_doc.file_url,
				total_rows=len(df),
				import_type="Other",
				field_mapping=normalized_mapping,  # mapping chiều đúng
				batch_size=batch_size,
				validate_only=validate_only,
				skip_duplicates=skip_duplicates
			)
			session.add_simple_log_talent("info", f"Import session started for {len(df)} rows")
			print(">>>>>>>>>>>>>>>>>>> 6.1.")
			session.update_status("Processing")

			print(">>>>>>>>>>>>>>>>>>> 7.")
			# 7️⃣ Chạy chế độ validate hoặc insert
			if validate_only:
				print(">>>>>>>>>>>>>>>>>>> 8.")
				results = validate_import_data_with_session_talent(df, normalized_mapping, session)
			else:
				print(">>>>>>>>>>>>>>>>>>> 9.")
				results = process_import_data_with_session_talent(df, normalized_mapping, session, batch_size, segment_id)

			print(">>>>>>>>>>>>>>>>>>> 10.")
			# 9️⃣ Cập nhật trạng thái session
			if results.get("failed", 0) == 0:
				session.update_status("Completed")
			else:
				session.update_status("Completed")
				session.error_summary = f"{results['failed']} rows failed during import"

			return {
				**results,
				"session_id": session.name
			}

		finally:
			# 🔟 Xoá file tạm
			try:
				frappe.delete_doc("File", file_doc.name, ignore_permissions=True)
			except Exception as cleanup_err:
				print(">>>>>>>>>>>>>>>>>>> 10. Lỗi khi xoá file tạm:", cleanup_err)

	except Exception as e:
		frappe.log_error(frappe.get_traceback(), "Candidate Import Error")
		frappe.throw(_("Import failed: {0}").format(str(e)))

# Hàm chung
@frappe.whitelist()
def get_import_sessions():
	"""Get list of import sessions with summary"""
	try:
		sessions = frappe.get_all(
			"Mira Importsession",
			fields=[
				"name", "file_name", "created_by", "started_at", "completed_at",
				"status", "total_rows", "success_count", "failed_count", 
				"import_type", "selected_job_opening", "validate_only", "error_summary"
			],
			order_by="started_at desc",
			limit=50
		)
		
		# Add retry information
		for session in sessions:
			# Check if this is a retry session
			if session.get("file_name", "").startswith("RETRY_"):
				session["is_retry"] = True
				# Get parent session if available
				parent_session = frappe.db.get_value("Mira Importsession", 
					{"name": session.name}, "parent_session")
				if parent_session:
					session["parent_session"] = parent_session
			else:
				session["is_retry"] = False
			
			# Count failed rows for retry possibility
			if session.status in ["Completed", "Failed"]:
				failed_count = 0
				if session.error_summary:
					try:
						all_logs = json.loads(session.error_summary)
						failed_count = len([log for log in all_logs if log.get("status") == "error"])
					except:
						pass
				session["can_retry"] = failed_count > 0
				session["retry_job_openings"] = failed_count
			else:
				session["can_retry"] = False
				session["retry_job_openings"] = 0
		
		return sessions
		
	except Exception as e:
		frappe.log_error(f"Error getting import sessions: {str(e)}", "Get Import Sessions Error")
		return []

@frappe.whitelist()
def delete_import_session(session_id: str):
	"""Delete an import session and its logs"""
	try:
		session = frappe.get_doc("Mira ImportSession", session_id)
		
		if session.status == "Processing":
			frappe.throw(_("Cannot delete a session that is currently processing"))
		
		# Delete associated logs first
		frappe.db.delete("ImportLog", {"parent": session_id})
		
		# Delete the session
		frappe.delete_doc("ImportSession", session_id, ignore_permissions=True)
		frappe.db.commit()
		
		return {
			"message": _("Import session deleted successfully")
		}
		
	except Exception as e:
		frappe.log_error(f"Error deleting import session: {str(e)}", "Delete Import Session Error")
		frappe.throw(_("Failed to delete session: {0}").format(str(e)))

@frappe.whitelist()
def export_failed_rows(session_id: str):
	"""Export failed rows to Excel for manual correction"""
	try:
		session = frappe.get_doc("Mira Importsession", session_id)
		
		# Get failed logs with details
		failed_logs = []
		if session.error_summary:
			try:
				all_logs = json.loads(session.error_summary)
				failed_logs = [log for log in all_logs if log.get("status") == "error"]
			except:
				pass
		
		if not failed_logs:
			frappe.throw(_("No failed rows found to export"))
		
		# Try to reconstruct the original data
		original_file = frappe.db.get_value("File", {"file_url": session.file_url}, "name")
		if original_file:
			try:
				file_doc = frappe.get_doc("File", original_file)
				df = read_file_to_dataframe(file_doc.get_full_path(), session.file_name)
				
				# Extract failed rows
				failed_row_numbers = [log["row_number"] - 1 for log in failed_logs]
				failed_df = df.iloc[failed_row_numbers].copy()
				
				# Add error information
				error_messages = [log.get("message", "") for log in failed_logs]
				failed_df.insert(0, 'Row_Number', [log["row_number"] for log in failed_logs])
				failed_df.insert(1, 'Error_Message', error_messages)
				
				# Create Excel file
				import tempfile
				with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
					with pd.ExcelWriter(tmp.name, engine='openpyxl') as writer:
						failed_df.to_excel(writer, sheet_name='Failed Rows', index=False)
						
						# Style the worksheet
						from openpyxl.styles import Font, PatternFill
						workbook = writer.book
						worksheet = writer.sheets['Failed Rows']
						
						# Style error columns
						error_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
						header_font = Font(bold=True)
						
						# Apply styling
						for col in range(1, 3):  # Row_Number and Error_Message columns
							for row in range(2, len(failed_df) + 2):
								worksheet.cell(row=row, column=col).fill = error_fill
						
						# Style headers
						for col in range(1, len(failed_df.columns) + 1):
							worksheet.cell(row=1, column=col).font = header_font
					
					# Read the file
					with open(tmp.name, 'rb') as f:
						content = f.read()
					
					# Clean up
					os.unlink(tmp.name)
				
				# Create Frappe file
				file_doc = frappe.get_doc({
					"doctype": "File",
					"file_name": f"failed_rows_{session_id}_{frappe.utils.now_datetime().strftime('%Y%m%d_%H%M%S')}.xlsx",
					"content": content,
					"is_private": 0
				})
				file_doc.save(ignore_permissions=True)
				
				return {
					"file_url": file_doc.file_url,
					"file_name": file_doc.file_name,
					"failed_count": len(failed_logs)
				}
				
			except Exception as e:
				frappe.log_error(f"Error creating failed rows export: {str(e)}")
				# Fallback to simple error list
				pass
		
		# Fallback: Create simple error summary
		error_data = []
		for log in failed_logs:
			error_data.append({
				'Row_Number': log.get("row_number", 0),
				'Job_Title': log.get("job_title", "Unknown"),
				'Error_Message': log.get("message", "Unknown error")
			})
		
		df = pd.DataFrame(error_data)
		
		# Create Excel file
		import tempfile
		with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
			df.to_excel(tmp.name, index=False)
			
			with open(tmp.name, 'rb') as f:
				content = f.read()
			
			os.unlink(tmp.name)
		
		# Create Frappe file
		file_doc = frappe.get_doc({
			"doctype": "File",
			"file_name": f"import_errors_{session_id}_{frappe.utils.now_datetime().strftime('%Y%m%d_%H%M%S')}.xlsx",
			"content": content,
			"is_private": 0
		})
		file_doc.save(ignore_permissions=True)
		
		return {
			"file_url": file_doc.file_url,
			"file_name": file_doc.file_name,
			"failed_count": len(failed_logs)
		}
		
	except Exception as e:
		frappe.log_error(frappe.get_traceback(), "Export Failed Rows Error")
		frappe.throw(_("Failed to export failed rows: {0}").format(str(e)))


# ===== HELPER FUNCTIONS FOR JOB OPENING IMPORT =====

#job 1
def create_import_session(file_name, file_url, total_rows, import_type="Job Opening",
						 selected_job_opening=None, field_mapping=None, batch_size=100,
						 validate_only=False, skip_duplicates=True, **kwargs):
	"""Create a new import session (Mira Importsession)"""
	session = frappe.new_doc("Mira Importsession")
	session.file_name = file_name
	session.file_url = file_url
	session.total_rows = total_rows
	session.import_type = import_type
	session.selected_job_opening = selected_job_opening
	session.field_mapping = json.dumps(field_mapping) if field_mapping else None
	session.status = "Pending"
	session.batch_size = batch_size
	session.validate_only = 1 if validate_only else 0
	session.skip_duplicates = 1 if skip_duplicates else 0
	session.created_by = frappe.session.user
	session.started_at = frappe.utils.now()
	# Set any additional kwargs
	for key, value in kwargs.items():
		if hasattr(session, key):
			setattr(session, key, value)
	session.insert(ignore_permissions=True)
	frappe.db.commit()
	return session

#can 1
def create_import_session_candidate(file_name, file_url, total_rows, import_type="Mira Candidate",
						 selected_job_opening=None, field_mapping=None, batch_size=100,
						 validate_only=False, skip_duplicates=True, **kwargs):
	"""Create a new import session (Mira Importsession)"""
	session = frappe.new_doc("Mira Importsession")
	session.file_name = file_name
	session.file_url = file_url
	session.total_rows = total_rows
	session.import_type = import_type
	session.selected_job_opening = selected_job_opening
	session.field_mapping = json.dumps(field_mapping) if field_mapping else None
	session.status = "Pending"
	session.batch_size = batch_size
	session.validate_only = 1 if validate_only else 0
	session.skip_duplicates = 1 if skip_duplicates else 0
	session.created_by = frappe.session.user
	session.started_at = frappe.utils.now()
	# Set any additional kwargs
	for key, value in kwargs.items():
		if hasattr(session, key):
			setattr(session, key, value)
	session.insert(ignore_permissions=True)
	frappe.db.commit()
	return session

#talent 1
def create_import_session_talennt(file_name, file_url, total_rows, import_type="Other", field_mapping=None, batch_size=100,
                         validate_only=False, skip_duplicates=True, **kwargs):
    """Create a new import session (Mira Importsession)"""
    session = frappe.new_doc("Mira Importsession")
    session.file_name = file_name
    session.file_url = file_url
    session.total_rows = total_rows
    session.import_type = import_type  # This will be "Other" by default
    session.field_mapping = json.dumps(field_mapping) if field_mapping else None
    session.status = "Pending"
    session.batch_size = batch_size
    session.validate_only = 1 if validate_only else 0
    session.skip_duplicates = 1 if skip_duplicates else 0
    session.created_by = frappe.session.user
    session.started_at = frappe.utils.now()
    
    # Set any additional kwargs
    for key, value in kwargs.items():
        if hasattr(session, key):
            setattr(session, key, value)
    
    session.insert(ignore_permissions=True)
    frappe.db.commit()
    return session

def read_file_to_dataframe(file_path: str, filename: str) -> pd.DataFrame:
	"""Read file into DataFrame with proper error handling"""
	ext = filename.split('.')[-1].lower()
	try:
		if ext in ['xlsx', 'xls']:
			# 👉 Force all cells to be read as strings to keep leading zeros (e.g. phone numbers)
			df = pd.read_excel(file_path, dtype=str, engine='openpyxl' if ext == 'xlsx' else None)
		elif ext == 'csv':
			# Try different encodings
			encodings = ['utf-8', 'latin-1', 'cp1252']
			df = None
			for encoding in encodings:
				try:
					# 👉 Force all columns to string to keep leading zeros
					df = pd.read_csv(file_path, encoding=encoding, dtype=str, keep_default_na=False)
					break
				except (UnicodeDecodeError, UnicodeError):
					continue
			if df is None:
				frappe.throw(_("Cannot read CSV file with any supported encoding"))
		else:
			frappe.throw(_("Unsupported file format"))
		# Basic validation
		if df.empty:
			frappe.throw(_("File is empty"))
		# Clean data
		df = df.dropna(how='all')
		df.columns = [str(col).strip() for col in df.columns]
		
		# ✅ Convert all values to string (avoid float/int converting again)
		df = df.applymap(lambda x: str(x).strip() if pd.notna(x) else "")
		
		return df
	except Exception as e:
		frappe.throw(_("Error reading file: {0}").format(str(e)))


def map_row_data(row: pd.Series, mapping: Dict) -> Dict:
	"""Map row data according to field mapping (Excel Header -> DocField)"""
	job_opening_data: Dict[str, Any] = {}
	for field_name, column_name in mapping.items():
		if column_name and column_name in row.index:
			value = row[column_name]
			if pd.notna(value) and str(value).strip():
				job_opening_data[field_name] = str(value).strip()
	return job_opening_data

def map_row_data_candidate(row: pd.Series, mapping) -> Dict:
	candidate_data: Dict[str, Any] = {}

	try:
		if hasattr(mapping, "field_mapping"):
			try:
				mapping_raw = getattr(mapping, "field_mapping") or "{}"
				mapping = json.loads(mapping_raw)
			except Exception as e:
				mapping = {}

	except Exception as e:
		print(">>>>>>>>>>>>>>>>>>>>>>>>>7.3.3", e)

	try:
		# lowercase map for column matching
		row_lower_map = {col.lower(): col for col in row.index}

		for field_name, column_name in mapping.items():
			col_lower = column_name.lower()
			if col_lower in row_lower_map:
				actual_col = row_lower_map[col_lower]
				value = row[actual_col]
				if pd.notna(value) and str(value).strip():
					candidate_data[field_name] = str(value).strip()
	except Exception as e:
		print(">>>>>>>>>>>>>>>>>>> 7.3.7", e)

	return candidate_data

def map_row_data_talent(row: pd.Series, mapping) -> Dict:
	talent_data: Dict[str, Any] = {}

	try:
		if hasattr(mapping, "field_mapping"):
			try:
				mapping_raw = getattr(mapping, "field_mapping") or "{}"
				mapping = json.loads(mapping_raw)
			except Exception as e:
				mapping = {}

	except Exception as e:
		print(">>>>>>>>>>>>>>>>>>>>>>>>>7.3.3", e)

	try:
		# lowercase map for column matching
		row_lower_map = {col.lower(): col for col in row.index}

		for field_name, column_name in mapping.items():
			col_lower = column_name.lower()
			if col_lower in row_lower_map:
				actual_col = row_lower_map[col_lower]
				value = row[actual_col]
				if pd.notna(value) and str(value).strip():
					talent_data[field_name] = str(value).strip()
	except Exception as e:
		print(">>>>>>>>>>>>>>>>>>> 7.3.7", e)

	return talent_data

# check trùng của job
def get_existing_job_titles() -> Set[str]:
	"""Get existing job titles to prevent duplicates within this import"""
	return set([
		title[0] for title in frappe.db.sql(
			"SELECT job_title FROM `tabJobOpening` WHERE job_title IS NOT NULL AND job_title != ''"
		)
	])

# check trùng của candidate
def get_existing_emails() -> Set[str]:
	"""Get all existing candidate emails from Mira Candidate to prevent duplicates"""
	return set([
		email[0] for email in frappe.db.sql(
			"SELECT email FROM `tabMira Candidate` WHERE email IS NOT NULL AND email != ''"
		)
	])

# check trùng của talent
def get_existing_talent_emails() -> Set[str]:
	"""Get all existing talent emails from Mira Talent to prevent duplicates"""
	return set([
		email[0].lower() for email in frappe.db.sql(
			"SELECT email FROM `tabMira Talent` WHERE email IS NOT NULL AND email != ''"
		)
	])

# của candidate 1
def validate_and_process_candidate(candidate_data: Dict, existing_emails: Set = None) -> Dict:
	"""Enhanced validation with better error messages and field processing"""
	
	processed = {}
	errors = []
	# Required field validation
	required_fields = {
		"full_name": _("Full Name"),
		"email": _("Email")
	}
	
	for field, label in required_fields.items():
		value = candidate_data.get(field, "")
		if not value or not str(value).strip():
			err_msg = f"Missing required field: {label}"
			errors.append(err_msg)
		else:
			processed_value = str(value).strip()
			processed[field] = processed_value
	# Email validation and duplicate check
	email = str(candidate_data.get("email", "")).strip().lower()
	
	if email:
		if not validate_email(email):
			err_msg = f"Invalid email format: {email}"
			errors.append(err_msg)
		elif existing_emails and email in existing_emails:
			err_msg = f"Email already exists: {email}"
			errors.append(err_msg)
		else:
			processed["email"] = email
			if existing_emails is not None:
				existing_emails.add(email)
	else:
		print("- Email is empty or missing")
	
	# Process other fields
	field_processors = {
		"full_name": lambda x: process_text_field(x, 140),
		"phone": process_phone_number,
		"avatar": lambda x: process_text_field(x, 255),
		"headline": lambda x: process_text_field(x, 140),
		"source": lambda x: process_text_field(x, 140),
		"cv_original_url": lambda x: process_text_field(x, 500),
		"skills": lambda x: process_text_field(x, 1000),
		"ai_summary": lambda x: process_text_field(x, 10000)  # For long text
	}
	
	for field, processor in field_processors.items():
		if field in candidate_data and candidate_data[field]:
			try:
				result = processor(candidate_data[field])
				if result is not None:
					processed[field] = result
			except Exception as e:
				errors.append(_("Error processing {0}: {1}").format(field, str(e)))
	
	# Set default status if not provided
	if "status" not in processed:
		processed["status"] = "NEW"
	
	if errors:
		raise frappe.ValidationError("; ".join(errors))
	
	return processed

# của talent 1
def validate_and_process_talent(talent_data: Dict, existing_emails: Set = None) -> Dict:
	"""Enhanced validation with better error messages and field processing"""
	
	processed = {}
	errors = []
	# Required field validation
	required_fields = {
		"full_name": _("Full Name"),
		"email": _("Email")
	}
	
	for field, label in required_fields.items():
		value = talent_data.get(field, "")
		if not value or not str(value).strip():
			err_msg = f"Missing required field: {label}"
			errors.append(err_msg)
		elif field == "email":
			# Email will be processed separately with lowercase and duplicate check
			continue
		else:
			processed_value = str(value).strip()
			processed[field] = processed_value
	
	# Email validation and duplicate check
	email = str(talent_data.get("email", "")).strip().lower()
	
	if email:
		if not validate_email(email):
			err_msg = f"Invalid email format: {email}"
			errors.append(err_msg)
		elif existing_emails and email in existing_emails:
			err_msg = f"Email already exists: {email}"
			errors.append(err_msg)
		else:
			processed["email"] = email
			if existing_emails is not None:
				existing_emails.add(email)
	else:
		errors.append("Missing required field: Email")
	
	# Process other fields
	field_processors = {
		"full_name": lambda x: process_text_field(x, 140),
		"phone": process_phone_number,
		"source": process_talent_source,
		"skills": lambda x: process_text_field(x, 1000),
		"crm_status": process_talent_crm_status,
		"gender": lambda x: process_text_field(x, 50),
		"date_of_birth": process_date_field,
		"linkedin_profile": lambda x: process_text_field(x, 255),
		"facebook_profile": lambda x: process_text_field(x, 255),
		"zalo_profile": lambda x: process_text_field(x, 255),
		"current_city": lambda x: process_text_field(x, 255),
		"desired_role": lambda x: process_text_field(x, 255),
		"domain_expertise": lambda x: process_text_field(x, 1000),
		"hard_skills": lambda x: process_text_field(x, 1000),
		"soft_skills": lambda x: process_text_field(x, 1000),
		"total_years_of_experience": lambda x: process_numeric_field(x),
		"latest_company": lambda x: process_text_field(x, 255),
		"highest_education": lambda x: process_text_field(x, 255),
		"current_salary": lambda x: process_numeric_field(x),
		"expected_salary": lambda x: process_numeric_field(x),
		"preferred_work_model": lambda x: process_text_field(x, 50),
		"availability_date": process_date_field,
		"recruiter_owner_id": process_owner_id,
		"recruitment_readiness": lambda x: process_text_field(x, 50),
		"last_interaction_date": process_date_field,
		"internal_rating": lambda x: process_text_field(x, 10),
		"priority_level": lambda x: process_text_field(x, 50),
		"interaction_notes": lambda x: process_text_field(x, 2000),
		"tags": lambda x: process_text_field(x, 1000),
		"cultural_fit": lambda x: process_text_field(x, 50),
		"notes": lambda x: process_text_field(x, 2000),
		"current_status": lambda x: process_text_field(x, 50),
		"latest_title": lambda x: process_text_field(x, 140),
	}
	
	for field, processor in field_processors.items():
		if field in talent_data and talent_data[field]:
			try:
				original_value = talent_data[field]
				result = processor(original_value)
				if result is not None:
					processed[field] = result
					print(f"Field '{field}': '{original_value}' -> '{result}'")
				else:
					print(f"Field '{field}': '{original_value}' -> None (validation failed)")
			except Exception as e:
				error_msg = _("Error processing {0}: {1}").format(field, str(e))
				errors.append(error_msg)
				print(f"Error processing field '{field}': {e}")
	
	# Set default values if not provided
	if "status" not in processed:
		processed["status"] = "NEW"
	
	# Set default source to NEW if not provided
	if "source" not in processed:
		processed["source"] = "NEW"
	
	# Set default crm_status to New if not provided
	if "crm_status" not in processed:
		processed["crm_status"] = "New"
	
	if errors:
		raise frappe.ValidationError("; ".join(errors))
	
	return processed

#của talent 2
def validate_import_data_with_session_talent(df: pd.DataFrame, mapping: Dict, session=None) -> Dict:
	"""Validate talent import data with session tracking
	"""   
	results = {
		"success": 0,
		"failed": 0,
		"total": len(df),
		"logs": []
	}
	# Get existing emails using the helper function for Talent
	existing_emails = get_existing_talent_emails()
	seen_emails = set()  # Track emails in current import to catch duplicates
	for idx, row in df.iterrows():
		start_time = time.time()
		row_number = idx + 1

		try:
			talent_data = map_row_data_talent(row, mapping)

			processed_data = validate_and_process_talent(talent_data, existing_emails)
			
			email = processed_data.get("email", "").lower()
			if email and email in seen_emails:
				raise frappe.ValidationError(_("Duplicate email in import file: {0}").format(email))
			if email:
				seen_emails.add(email)
			
			processing_time = time.time() - start_time

			if session:
				session.mark_success_talent(
					row_number=row_number,
					full_name=processed_data.get("full_name", "Unknown"),
					processing_time=processing_time
				)
			results["success"] += 1
			results["logs"].append({
				"row_number": row_number,
				"full_name": processed_data.get("full_name", "Unknown"),
				"status": "success",
				"message": _("Validation passed")
			})
				
		except Exception as e:
			processing_time = time.time() - start_time
			error_msg = str(e)

			if session:
				session.mark_failure_talent(
					row_number=row_number,
					full_name=talent_data.get("full_name", "Unknown") if 'talent_data' in locals() else "Unknown",
					error_message=error_msg,
					processing_time=processing_time
				)

			results["failed"] += 1
			results["logs"].append({
				"row_number": row_number,
				"candidate_name": talent_data.get("full_name", "Unknown") if 'candidate_data' in locals() else "Unknown",
				"status": "error",
				"message": error_msg
			})
		
		# Update session progress periodically
		if session and idx % 10 == 0:
			session.save(ignore_permissions=True)
			frappe.db.commit()
	
	if session:
		session.save(ignore_permissions=True)
		frappe.db.commit()
	
	return results

#của candidate 2
def validate_import_data_with_session_candidate(df: pd.DataFrame,job_opening, mapping: Dict, session=None) -> Dict:
	"""Validate candidate import data with session tracking
	"""   
	results = {
		"success": 0,
		"failed": 0,
		"total": len(df),
		"logs": []
	}
	# Get existing emails using the helper function
	existing_emails = get_existing_emails()
	seen_emails = set()  # Track emails in current import to catch duplicates
	for idx, row in df.iterrows():
		start_time = time.time()
		row_number = idx + 1

		try:
			candidate_data = map_row_data_candidate(row, mapping)

			processed_data = validate_and_process_candidate(candidate_data, existing_emails)
			
			email = processed_data.get("email", "").lower()
			if email and email in seen_emails:
				raise frappe.ValidationError(_("Duplicate email in import file: {0}").format(email))
			if email:
				seen_emails.add(email)
			
			processing_time = time.time() - start_time

			if session:
				session.mark_success_candidate(
					row_number=row_number,
					full_name=processed_data.get("full_name", "Unknown"),
					processing_time=processing_time
				)
			results["success"] += 1
			results["logs"].append({
				"row_number": row_number,
				"full_name": processed_data.get("full_name", "Unknown"),
				"status": "success",
				"message": _("Validation passed")
			})
				
		except Exception as e:
			processing_time = time.time() - start_time
			error_msg = str(e)

			if session:
				session.mark_failure_candidate(
					row_number=row_number,
					full_name=candidate_data.get("full_name", "Unknown") if 'candidate_data' in locals() else "Unknown",
					error_message=error_msg,
					processing_time=processing_time
				)

			results["failed"] += 1
			results["logs"].append({
				"row_number": row_number,
				"candidate_name": candidate_data.get("full_name", "Unknown") if 'candidate_data' in locals() else "Unknown",
				"status": "error",
				"message": error_msg
			})
		
		# Update session progress periodically
		if session and idx % 10 == 0:
			session.save(ignore_permissions=True)
			frappe.db.commit()
	
	if session:
		session.save(ignore_permissions=True)
		frappe.db.commit()
	
	return results

# của job 1
def validate_and_process_job_opening(job_opening_data: Dict, existing_titles: Set = None) -> Dict:
	"""Validate and normalize JobOpening data"""
	processed: Dict[str, Any] = {}
	errors: List[str] = []
	# Required field(s)
	required_fields = {
		"job_title": _("Job Title")
	}
	for field, label in required_fields.items():
		value = (job_opening_data.get(field) or "").strip()
		if not value:
			errors.append(_("Missing required field: {0}").format(label))
	# Duplicate title check within system
	job_title = (job_opening_data.get("job_title") or "").strip()
	if job_title:
		if existing_titles and job_title in existing_titles:
			errors.append(_("Job title already exists: {0}").format(job_title))
		else:
			processed["job_title"] = job_title
			if existing_titles is not None:
				existing_titles.add(job_title)
	# Optional fields processing
	field_processors = {
		"description": lambda x: process_text_field(x, 2000),
		"requirements": lambda x: process_text_field(x, 2000),
		"benefits": lambda x: process_text_field(x, 2000),
		"department_name": lambda x: process_text_field(x, 140),
		"location_name": lambda x: process_text_field(x, 140),
		"number_of_openings": process_number_of_openings,
		"posting_date": process_date,
		"closing_date": process_date,
		"approval_status": process_approval_status,
		"owner_id": process_owner_id,
	}
	for field, processor in field_processors.items():
		if field in job_opening_data and job_opening_data[field]:
			try:
				result = processor(job_opening_data[field])
				if result is not None:
					processed[field] = result
			except Exception as e:
				errors.append(_("Error processing {0}: {1}").format(field, str(e)))
	# Defaults
	if "approval_status" not in processed:
		processed["approval_status"] = "Draft"
	if "posting_date" not in processed:
		processed["posting_date"] = today()
	if errors:
		raise frappe.ValidationError("; ".join(errors))
	return processed

# của job 2
def validate_import_data_with_session(df: pd.DataFrame, mapping: Dict, selected_job_opening: str = None, session=None) -> Dict:
	"""Validate import data with session tracking (no insert)"""
	results = {"success": 0, "failed": 0, "total": len(df), "logs": []}
	existing_titles = get_existing_job_titles()
	seen_titles: Set[str] = set()
	for idx, row in df.iterrows():
		start_time = time.time()
		row_number = idx + 1
		try:
			job_opening_data = map_row_data(row, mapping)
			processed_data = validate_and_process_job_opening(job_opening_data, existing_titles)
			title = (processed_data.get("job_title") or "").strip()
			if title and title in seen_titles:
				raise frappe.ValidationError(_("Duplicate job title in import file: {0}").format(title))
			if title:
				seen_titles.add(title)
			processing_time = time.time() - start_time
			if session:
				session.mark_success(
					row_number=row_number,
					job_title=processed_data.get("job_title", "Unknown"),
					processing_time=processing_time
				)
			results["success"] += 1
			results["logs"].append({
				"row_number": row_number,
				"job_title": processed_data.get("job_title", "Unknown"),
				"status": "success",
				"message": _("Validation passed")
			})
		except Exception as e:
			processing_time = time.time() - start_time
			error_msg = str(e)
			if session:
				session.mark_failure(
					row_number=row_number,
					job_title=job_opening_data.get("job_title", "Unknown") if 'job_opening_data' in locals() else "Unknown",
					error_message=error_msg,
					processing_time=processing_time
				)
			results["failed"] += 1
			results["logs"].append({
				"row_number": row_number,
				"job_title": job_opening_data.get("job_title", "Unknown") if 'job_opening_data' in locals() else "Unknown",
				"status": "error",
				"message": error_msg
			})
		# Periodic save
		if session and idx % 10 == 0:
			session.save(ignore_permissions=True)
			frappe.db.commit()
	if session:
		session.save(ignore_permissions=True)
		frappe.db.commit()
	return results

#của job 3
def process_import_data_with_session(df: pd.DataFrame, mapping: Dict, selected_job_opening: str, session, batch_size: int = 100) -> Dict:
	"""Process and insert import data with session tracking"""
	results = {"success": 0, "failed": 0, "total": len(df), "logs": []}
	existing_titles = get_existing_job_titles()
	for start_idx in range(0, len(df), batch_size):
		end_idx = min(start_idx + batch_size, len(df))
		batch_df = df.iloc[start_idx:end_idx]
		batch_results = process_batch_with_session(batch_df, mapping, selected_job_opening, existing_titles, start_idx, session)
		results["success"] += batch_results["success"]
		results["failed"] += batch_results["failed"]
		results["logs"].extend(batch_results["logs"])
		# Commit and publish progress
		session.save(ignore_permissions=True)
		frappe.db.commit()
		frappe.publish_realtime(
			'import_progress',
			{
				'session_id': session.name,
				'processed': end_idx,
				'total': len(df),
				'success': results["success"],
				'failed': results["failed"],
				'progress_percentage': (end_idx / len(df)) * 100
			}
		)
		time.sleep(0.1)
	return results

#job 4 
def process_batch_with_session(df: pd.DataFrame, mapping: Dict, selected_job_opening: str,
								existing_titles: Set, start_offset: int = 0, session=None) -> Dict:
	"""Process a single batch (insert JobOpening docs)"""
	batch_results = {"success": 0, "failed": 0, "logs": [], "new_titles": []}
	for idx, row in df.iterrows():
		actual_row_num = start_offset + idx + 1
		start_time = time.time()
		try:
			job_opening_data = map_row_data(row, mapping)
			processed_data = validate_and_process_job_opening(job_opening_data, existing_titles)
			doc = frappe.get_doc({"doctype": "Mira Job Opening", **processed_data})
			doc.insert(ignore_permissions=True)
			processing_time = time.time() - start_time
			title = processed_data.get("job_title")
			if title:
				batch_results["new_titles"].append(title)
			if session:
				session.mark_success(
					row_number=actual_row_num,
					job_title=processed_data.get("job_title", "Unknown"),
					document_name=doc.name,
					processing_time=processing_time
				)
			batch_results["success"] += 1
			batch_results["logs"].append({
				"row_number": actual_row_num,
				"job_title": processed_data.get("job_title", "Unknown"),
				"status": "success",
				"message": _("Successfully created: {0}").format(doc.name),
				"document_name": doc.name
			})
		except Exception as e:
			processing_time = time.time() - start_time
			error_msg = str(e)
			if session:
				session.mark_failure(
					row_number=actual_row_num,
					job_title=job_opening_data.get("job_title", "Unknown") if 'job_opening_data' in locals() else "Unknown",
					error_message=error_msg,
					processing_time=processing_time
				)
			batch_results["failed"] += 1
			batch_results["logs"].append({
				"row_number": actual_row_num,
				"job_title": job_opening_data.get("job_title", "Unknown") if 'job_opening_data' in locals() else "Unknown",
				"status": "error",
				"message": error_msg
			})
	return batch_results

#talent 3
def process_import_data_with_session_talent(df: pd.DataFrame, mapping: Dict, session, batch_size: int = 100, segment_id: str = None) -> Dict:
	"""Process and insert talent import data with session tracking"""
	results = {"success": 0, "failed": 0, "total": len(df), "logs": []}
	existing_emails = get_existing_talent_emails()
	for start_idx in range(0, len(df), batch_size):
		end_idx = min(start_idx + batch_size, len(df))
		batch_df = df.iloc[start_idx:end_idx]
		batch_results = process_batch_with_session_talent(
			batch_df, mapping, existing_emails, start_idx, session, segment_id
		)
		
		results["success"] += batch_results["success"]
		results["failed"] += batch_results["failed"]
		results["logs"].extend(batch_results["logs"])
		
		# Commit and publish progress
		session.save(ignore_permissions=True)
		frappe.db.commit()
		
		frappe.publish_realtime(
			'import_progress',
			{
				'session_id': session.name,
				'processed': end_idx,
				'total': len(df),
				'success': results["success"],
				'failed': results["failed"],
				'progress_percentage': (end_idx / len(df)) * 100
			}
		)
		time.sleep(0.1)
	return results

#can 3
def process_import_data_with_session_candidate(df: pd.DataFrame, mapping: Dict, selected_job_opening: str, session, batch_size: int = 100) -> Dict:
	"""Process and insert candidate import data with session tracking"""
	results = {"success": 0, "failed": 0, "total": len(df), "logs": []}
	existing_emails = get_existing_emails()
	for start_idx in range(0, len(df), batch_size):
		end_idx = min(start_idx + batch_size, len(df))
		batch_df = df.iloc[start_idx:end_idx]
		batch_results = process_batch_with_session_candidate(
			batch_df, mapping, selected_job_opening, existing_emails, start_idx, session
		)
		
		results["success"] += batch_results["success"]
		results["failed"] += batch_results["failed"]
		results["logs"].extend(batch_results["logs"])
		
		# Commit and publish progress
		session.save(ignore_permissions=True)
		frappe.db.commit()
		
		frappe.publish_realtime(
			'import_progress',
			{
				'session_id': session.name,
				'processed': end_idx,
				'total': len(df),
				'success': results["success"],
				'failed': results["failed"],
				'progress_percentage': (end_idx / len(df)) * 100
			}
		)
		time.sleep(0.1)
	return results

#talent 4
def process_batch_with_session_talent(df: pd.DataFrame, mapping: Dict, 
							  existing_emails: Set, start_offset: int = 0, session=None, segment_id: str = None) -> Dict:
	"""Process a single batch with session tracking for Mira Talent"""
	batch_results = {
		"success": 0,
		"failed": 0,
		"logs": [],
		"new_emails": []
	}
	
	for idx, row in df.iterrows():
		actual_row_num = start_offset + idx + 1
		start_time = time.time()
		
		try:
			# Map fields
			talent_data = map_row_data_talent(row, mapping)
			
			# Validate and process
			processed_data = validate_and_process_talent(talent_data, existing_emails)
			
			# Create talent document
			doc = frappe.get_doc({
				"doctype": "Mira Talent",
				**processed_data
			})
			doc.insert(ignore_permissions=True)
			
			# Create Mira Talent Pool record if segment_id is provided
			if segment_id:
				try:
					pool_doc = frappe.get_doc({
						"doctype": "Mira Talent Pool",
						"talent_id": doc.name,
						"segment_id": segment_id,
						"added_at": frappe.utils.now_datetime(),
						"added_by": frappe.session.user
					})
					pool_doc.insert(ignore_permissions=True)
				except Exception as pool_error:
					# Log error but don't fail the talent creation
					frappe.log_error(
						title="Talent Pool Creation Error",
						message=f"Failed to add talent {doc.name} to segment {segment_id}: {str(pool_error)}"
					)
			
			processing_time = time.time() - start_time
			
			# Track new email
			email = processed_data.get("email")
			if email:
				batch_results["new_emails"].append(email.lower())
			
			if session:
				session.mark_success_talent(
					row_number=actual_row_num,
					full_name=processed_data.get("full_name", "Unknown"),
					document_name=doc.name,
					processing_time=processing_time
				)
			
			batch_results["success"] += 1
			batch_results["logs"].append({
				"row_number": actual_row_num,
				"candidate_name": processed_data.get("full_name", "Unknown"),
				"status": "success",
				"message": _("Successfully created: {0}").format(doc.name),
				"document_name": doc.name
			})
			
		except Exception as e:
			processing_time = time.time() - start_time
			error_msg = str(e)
			
			if session:
				session.mark_failure_talent(
					row_number=actual_row_num,
					full_name=talent_data.get("full_name", "Unknown") if 'talent_data' in locals() else "Unknown",
					error_message=error_msg,
					processing_time=processing_time
				)
			
			batch_results["failed"] += 1
			batch_results["logs"].append({
				"row_number": actual_row_num,
				"talent_name": talent_data.get("full_name", "Unknown") if 'talent_data' in locals() else "Unknown",
				"status": "error",
				"message": error_msg
			})
	
	return batch_results

#can 4
def process_batch_with_session_candidate(df: pd.DataFrame, mapping: Dict, job_opening: str, 
							  existing_emails: Set, start_offset: int = 0, session=None) -> Dict:
	"""Process a single batch with session tracking for Mira_Candidate"""
	batch_results = {
		"success": 0,
		"failed": 0,
		"logs": [],
		"new_emails": []
	}
	
	for idx, row in df.iterrows():
		actual_row_num = start_offset + idx + 1
		start_time = time.time()
		
		try:
			# Map fields
			candidate_data = map_row_data(row, mapping)
			if job_opening:
				candidate_data["job_opening"] = job_opening
			
			# Validate and process
			processed_data = validate_and_process_candidate(candidate_data, existing_emails)
			
			# Create candidate document
			doc = frappe.get_doc({
				"doctype": "Mira Candidate",
				**processed_data
			})
			doc.insert(ignore_permissions=True)
			
			processing_time = time.time() - start_time
			
			# Track new email
			email = processed_data.get("email")
			if email:
				batch_results["new_emails"].append(email.lower())
			
			if session:
				session.mark_success_candidate(
					row_number=actual_row_num,
					full_name=processed_data.get("full_name", "Unknown"),
					document_name=doc.name,
					processing_time=processing_time
				)
			
			batch_results["success"] += 1
			batch_results["logs"].append({
				"row_number": actual_row_num,
				"talent_name": processed_data.get("full_name", "Unknown"),
				"status": "success",
				"message": _("Successfully created: {0}").format(doc.name),
				"document_name": doc.name
			})
			
		except Exception as e:
			processing_time = time.time() - start_time
			error_msg = str(e)
			
			if session:
				session.mark_failure_candidate(
					row_number=actual_row_num,
					full_name=candidate_data.get("full_name", "Unknown") if 'candidate_data' in locals() else "Unknown",
					error_message=error_msg,
					processing_time=processing_time
				)
			
			batch_results["failed"] += 1
			batch_results["logs"].append({
				"row_number": actual_row_num,
				"candidate_name": candidate_data.get("full_name", "Unknown") if 'candidate_data' in locals() else "Unknown",
				"status": "error",
				"message": error_msg
			})
	
	return batch_results



def process_text_field(value: Any, max_length: int = None) -> Optional[str]:
	"""Process text with optional max length"""
	if not value:
		return None
	text = str(value).strip()
	if not text:
		return None
	if max_length and len(text) > max_length:
		text = text[:max_length]
	return text


def process_number_of_openings(value: Any) -> Optional[int]:
	"""Validate and convert openings to int"""
	if value is None or value == "":
		return None
	try:
		num = int(float(str(value)))
		if num < 0:
			return None
		return num
	except (ValueError, TypeError):
		return None


def process_date(date_value: Any) -> Optional[str]:
	"""Parse date into YYYY-MM-DD (limited future range)"""
	if not date_value:
		return None
	if isinstance(date_value, datetime):
		return date_value.date().strftime('%Y-%m-%d')
	if isinstance(date_value, str):
		date_str = date_value.strip()
		if not date_str:
			return None
		date_formats = [
			"%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y",
			"%Y/%m/%d", "%d.%m.%Y", "%Y.%m.%d", "%B %d, %Y",
			"%d/%m/%y", "%m/%d/%y", "%d-%m-%y", "%y-%m-%d"
		]
		for fmt in date_formats:
			try:
				parsed_date = datetime.strptime(date_str, fmt).date()
				current_year = datetime.now().year
				if parsed_date.year < 2000 or parsed_date.year > current_year + 5:
					continue
				return parsed_date.strftime('%Y-%m-%d')
			except ValueError:
				continue
	return None


def process_approval_status(status: Any) -> Optional[str]:
	"""Normalize approval status"""
	if not status:
		return "Draft"
	status_str = str(status).strip()
	valid_statuses = ["Draft", "Pending Approval", "Approved", "Rejected"]
	if status_str in valid_statuses:
		return status_str
	return "Draft"


def process_owner_id(owner: Any) -> Optional[str]:
	"""Validate owner user id"""
	if not owner:
		return frappe.session.user
	owner_str = str(owner).strip()
	if frappe.db.exists("User", owner_str):
		return owner_str
	return frappe.session.user


def process_talent_source(source: Any) -> Optional[str]:
	"""Validate and normalize talent source field
	
	Args:
		source: The source value from import file
		
	Returns:
		Valid source value or raises ValidationError
		
	Raises:
		frappe.ValidationError: If source is not in allowed options
	"""
	if not source:
		return None
	
	source_str = str(source).strip()
	
	# Get valid source options from Mira Talent doctype
	meta = frappe.get_meta("Mira Talent")
	source_field = meta.get_field("source")
	
	if not source_field or not source_field.options:
		# If we can't get options, allow any value (fallback)
		return source_str
	
	# Parse options (newline-separated string)
	valid_sources = [opt.strip() for opt in source_field.options.split('\n') if opt.strip()]
	
	# Check if source value is in valid options (case-insensitive comparison)
	source_lower = source_str.lower()
	valid_sources_lower = {opt.lower(): opt for opt in valid_sources}
	
	if source_lower in valid_sources_lower:
		# Return the correctly cased version from doctype
		return valid_sources_lower[source_lower]
	else:
		# Source is not valid - raise error
		valid_options_str = ", ".join([f"'{opt}'" for opt in valid_sources if opt])
		raise frappe.ValidationError(
			_("Invalid source value: '{0}'. Allowed values are: {1}").format(
				source_str, valid_options_str
			)
		)


def process_talent_crm_status(crm_status: Any) -> Optional[str]:
	"""Validate and normalize talent crm_status field
	
	Args:
		crm_status: The crm_status value from import file
		
	Returns:
		Valid crm_status value or raises ValidationError
		
	Raises:
		frappe.ValidationError: If crm_status is not in allowed options
	"""
	if not crm_status:
		return None
	
	crm_status_str = str(crm_status).strip()
	
	# Get valid crm_status options from Mira Talent doctype
	meta = frappe.get_meta("Mira Talent")
	crm_status_field = meta.get_field("crm_status")
	
	if not crm_status_field or not crm_status_field.options:
		# If we can't get options, allow any value (fallback)
		return crm_status_str
	
	# Parse options (newline-separated string)
	valid_statuses = [opt.strip() for opt in crm_status_field.options.split('\n') if opt.strip()]
	
	# Check if crm_status value is in valid options (case-insensitive comparison)
	status_lower = crm_status_str.lower()
	valid_statuses_lower = {opt.lower(): opt for opt in valid_statuses}
	
	if status_lower in valid_statuses_lower:
		# Return the correctly cased version from doctype
		return valid_statuses_lower[status_lower]
	else:
		# CRM Status is not valid - raise error
		valid_options_str = ", ".join([f"'{opt}'" for opt in valid_statuses if opt])
		raise frappe.ValidationError(
			_("Invalid crm_status value: '{0}'. Allowed values are: {1}").format(
				crm_status_str, valid_options_str
			)
		)

#chung
@frappe.whitelist()
def cancel_import_session(session_id: str):
	"""Cancel an active import session"""
	try:
		session = frappe.get_doc("Mira ImportSession", session_id)
		
		if session.status not in ["Processing", "Pending"]:
			frappe.throw(_("Can only cancel sessions that are pending or in progress"))
		
		session.update_status("Cancelled")
		session.add_simple_log("warning", "Import session cancelled by user")
		
		# Note: This won't stop background jobs that are already running
		# You might need additional logic to handle job cancellation
		
		return {
			"message": _("Import session cancelled successfully"),
			"status": "Cancelled"
		}
		
	except Exception as e:
		frappe.log_error(f"Error cancelling import session: {str(e)}", "Cancel Import Error")
		frappe.throw(_("Failed to cancel import: {0}").format(str(e)))

@frappe.whitelist()
def get_retry_job_opening(session_id: str):
	"""Get list of failed rows that can be retried"""
	try:
		session = frappe.get_doc("Mira ImportSession", session_id)
		
		failed_logs = frappe.get_all(
			"Mira ImportLog",
			filters={
				"parent": session_id,
				"status": "Failed"
			},
			fields=["row_number", "job_title", "error_message", "processing_time"],
			order_by="row_number"
		)
		
		return {
			"session_name": session.name,
			"total_failed": len(failed_logs),
			"failed_rows": failed_logs,
			"can_retry": len(failed_logs) > 0 and session.status in ["Completed", "Failed"]
		}
		
	except frappe.DoesNotExistError:
		return {"error": "Session not found"}
	except Exception as e:
		frappe.log_error(f"Error getting retry candidates: {str(e)}", "Retry Candidates Error")
		return {"error": str(e)}

#chung
@frappe.whitelist()
def delete_import_session(session_id: str):
	"""Retry failed rows from an import session"""
	try:
		session = frappe.get_doc("Mira ImportSession", session_id)
		
		if session.status not in ["Completed", "Failed"]:
			frappe.throw(_("Cannot retry an import that is still in progress"))
		
		# Get failed import logs
		failed_logs = frappe.get_all(
			"Mira ImportLog",
			filters={
				"parent": session_id,
				"status": "Failed"
			},
			fields=["row_number", "job_title", "error_message"]
		)
		
		if not failed_logs:
			return {
				"retried": 0,
				"message": _("No failed rows to retry")
			}
		
		# Re-read the original file if it still exists
		original_file = frappe.db.get_value("File", {"file_url": session.file_url}, "name")
		if not original_file:
			frappe.throw(_("Original import file no longer exists. Cannot retry."))
		
		file_doc = frappe.get_doc("File", original_file)
		df = read_file_to_dataframe(file_doc.get_full_path(), session.file_name)
		
		# Get the field mapping
		mapping = json.loads(session.field_mapping) if session.field_mapping else {}
		if not mapping:
			frappe.throw(_("Field mapping not found in session"))
		
		# Create new retry session
		retry_session = create_import_session(
			file_name=f"RETRY_{session.file_name}",
			file_url=session.file_url,
			total_rows=len(failed_logs),
			import_type=session.import_type,
			job_opening=session.job_opening,
			field_mapping=mapping,
			batch_size=session.batch_size,
			validate_only=False,  # Always process for retry
			skip_duplicates=session.skip_duplicates,
			parent_session=session_id
		)
		
		retry_session.add_simple_log("info", f"Retry session started for {len(failed_logs)} failed rows from session {session_id}")
		retry_session.update_status("Processing")
		
		# Extract failed row numbers
		failed_row_numbers = [log.row_number - 1 for log in failed_logs]  # Convert to 0-based index
		failed_df = df.iloc[failed_row_numbers]
		
		# Process the failed rows
		results = process_import_data_with_session(
			failed_df, 
			mapping, 
			session.job_opening, 
			retry_session, 
			session.batch_size
		)
		
		# Update retry session status
		if results["failed"] == 0:
			retry_session.update_status("Completed")
			retry_session.add_simple_log("success", f"All {results['success']} rows processed successfully")
		else:
			retry_session.update_status("Completed")
			retry_session.error_summary = f"{results['failed']} rows still failed after retry"
			retry_session.add_simple_log("warning", f"{results['success']} succeeded, {results['failed']} still failed")
		
		return {
			"retried": len(failed_logs),
			"success": results["success"],
			"failed": results["failed"],
			"retry_session_id": retry_session.name,
			"message": _(f"Retry completed: {results['success']} successful, {results['failed']} still failed")
		}
		
	except Exception as e:
		frappe.log_error(frappe.get_traceback(), "Import Retry Error")
		frappe.throw(_("Retry failed: {0}").format(str(e)))

# chung
@frappe.whitelist()
def get_import_progress(session_id: str):
	"""Get import progress for a session"""
	try:
		session = frappe.get_doc("Mira ImportSession", session_id)
		summary = session.get_summary()
		summary["logs"] = session.get_logs(50)  # Get last 50 logs
		return summary
	except frappe.DoesNotExistError:
		return {"status": "Not Found"}
	except Exception as e:
		frappe.log_error(f"Error getting import progress: {str(e)}", "Import Progress Error")
		return {"status": "Error", "message": str(e)}


# def validate_and_get_candidate_source(source):
#     """Validate and return candidate source ID"""
#     if frappe.db.exists("CandidateSource", source):
#         return source
#     return None

# def validate_and_get_region(region):
#     """Validate and return region ID"""
#     if frappe.db.exists("Region", region):
#         return region
#     return None

# def process_gender(gender: Any) -> Optional[str]:
#     """Process gender with Vietnamese and English support"""
#     if not gender:
#         return None
	
#     gender_str = str(gender).lower().strip()
	
#     # Vietnamese mappings
#     male_variants = ['nam', 'male', 'm', 'boy', 'man', '1']
#     female_variants = ['nữ', 'nu', 'female', 'f', 'girl', 'woman', '0']
	
#     if gender_str in male_variants:
#         return 'Nam'
#     elif gender_str in female_variants:
#         return 'Nữ'
	
#     return None

def process_numeric_field(value: Any) -> Optional[float]:
	"""Process numeric field (for salary, experience years, etc.)"""
	if not value:
		return None
	
	try:
		# Handle string numbers with commas
		if isinstance(value, str):
			# Remove commas, spaces, and other formatting
			cleaned = re.sub(r'[,\s]', '', value.strip())
			if not cleaned:
				return None
			return float(cleaned)
		return float(value)
	except (ValueError, TypeError):
		return None

def process_date_field(date_value: Any) -> Optional[str]:
	"""Process date field with multiple format support"""
	if not date_value:
		return None
	
	# If already a date object
	if isinstance(date_value, (date, datetime)):
		return date_value.strftime('%Y-%m-%d')
	
	# If string, try to parse
	if isinstance(date_value, str):
		date_str = date_value.strip()
		if not date_str:
			return None
		
		date_formats = [
			"%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y",
			"%Y/%m/%d", "%d.%m.%Y", "%Y.%m.%d", "%B %d, %Y",
			"%d/%m/%y", "%m/%d/%y", "%d-%m-%y", "%y-%m-%d"
		]
		
		for fmt in date_formats:
			try:
				parsed_date = datetime.strptime(date_str, fmt).date()
				current_year = datetime.now().year
				# Reasonable date range validation
				if parsed_date.year < 1950 or parsed_date.year > current_year + 5:
					continue
				return parsed_date.strftime('%Y-%m-%d')
			except ValueError:
				continue
	
	return None

def process_phone_number(phone: Any) -> Optional[str]:
	"""Enhanced phone number processing"""
	if not phone:
		return None
	
	phone_str = str(phone).strip()
	if not phone_str:
		return None
	
	# Remove formatting but keep + for international numbers
	cleaned = re.sub(r'[^\d+]', '', phone_str)
	
	# Handle international format
	if cleaned.startswith('+'):
		cleaned = '+' + re.sub(r'[^\d]', '', cleaned[1:])
	else:
		cleaned = re.sub(r'[^\d]', '', cleaned)
	
	# Validate length (allow 9-15 digits for Vietnam phone numbers)
	digit_only = re.sub(r'[^\d]', '', cleaned)
	if len(digit_only) < 9 or len(digit_only) > 15:
		print(f"Phone validation failed: '{phone_str}' -> '{cleaned}' -> {len(digit_only)} digits")
		return None
	
	print(f"Phone processed: '{phone_str}' -> '{cleaned}'")
	return cleaned[:20]  # Limit total length


def validate_email(email: str) -> bool:
	"""Enhanced email validation"""
	if not email:
		return False
	
	# Basic format check
	pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
	if not re.match(pattern, email):
		return False
	
	# Additional checks
	if len(email) > 254:  # Email length limit
		return False
	
	local, domain = email.split('@')
	if len(local) > 64:  # Local part length limit
		return False
	
	return True