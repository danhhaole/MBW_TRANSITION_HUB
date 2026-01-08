import frappe
from frappe.model.document import Document
from frappe.utils import now, time_diff_in_seconds, flt
import json
from datetime import datetime
from frappe import _
from frappe.utils.file_manager import save_file
from frappe.utils import now, getdate, today
from datetime import datetime
import re, io, os
import json
import pandas as pd
import time
from typing import Dict, Any, Optional, Set


class TransitionHubImportSession(Document):
	def before_insert(self):
		"""Set default values before insert"""
		if not self.created_by:
			self.created_by = frappe.session.user
		if not self.started_at:
			self.started_at = now()
		if not self.status:
			self.status = "Pending"

		# Initialize logs storage
		self._temp_logs = []

	def before_save(self):
		"""Update calculated fields before save"""
		self.calculate_progress()
		self.estimate_time_remaining()

	def after_insert(self):
		"""Actions after document creation"""
		# Simple log without child table
		self.add_simple_log("info", f"Import session created for file: {self.file_name}")

	def calculate_progress(self):
		"""Calculate progress percentage"""
		if self.total_rows > 0:
			processed = (self.success_count or 0) + (self.failed_count or 0) + (self.duplicate_count or 0)
			self.progress_percentage = flt((processed / self.total_rows) * 100, 2)
		else:
			self.progress_percentage = 0

	def estimate_time_remaining(self):
		"""Estimate remaining time based on current progress"""
		if not self.started_at or self.status in ["Completed", "Failed", "Cancelled"]:
			self.estimated_time_remaining = ""
			return

		if self.progress_percentage <= 0:
			self.estimated_time_remaining = "Calculating..."
			return

		elapsed_time = time_diff_in_seconds(now(), self.started_at)
		if elapsed_time <= 0:
			self.estimated_time_remaining = "Calculating..."
			return

		# Calculate average time per row
		processed_rows = ((self.success_count or 0) + (self.failed_count or 0) + (self.duplicate_count or 0))
		if processed_rows <= 0:
			self.estimated_time_remaining = "Calculating..."
			return

		avg_time_per_row = elapsed_time / processed_rows
		remaining_rows = self.total_rows - processed_rows
		estimated_seconds = remaining_rows * avg_time_per_row

		if estimated_seconds <= 0:
			self.estimated_time_remaining = "Almost done"
		elif estimated_seconds < 60:
			self.estimated_time_remaining = f"{int(estimated_seconds)} seconds"
		elif estimated_seconds < 3600:
			minutes = int(estimated_seconds / 60)
			self.estimated_time_remaining = f"{minutes} minute(s)"
		else:
			hours = int(estimated_seconds / 3600)
			minutes = int((estimated_seconds % 3600) / 60)
			self.estimated_time_remaining = f"{hours}h {minutes}m"

	def add_simple_log(self, status, message, row_number=None, candidate_name=None,
						document_name=None, processing_time=None):
		"""Add a simple log entry without child table"""
		try:
			# Initialize temp logs if not exists
			if not hasattr(self, '_temp_logs'):
				self._temp_logs = []

			log_entry = {
				"timestamp": now(),
				"status": status,
				"message": message,
				"row_number": row_number,
				"candidate_name": candidate_name,
				"document_name": document_name,
				"processing_time": processing_time
			}

			self._temp_logs.append(log_entry)

			# Store in error_summary as JSON (keep last 100 entries)
			if len(self._temp_logs) > 100:
				self._temp_logs = self._temp_logs[-100:]

			# Update error_summary field with logs
			try:
				existing_logs = []
				if self.error_summary:
					existing_logs = json.loads(self.error_summary)

				existing_logs.append(log_entry)
				if len(existing_logs) > 100:
					existing_logs = existing_logs[-100:]

				self.error_summary = json.dumps(existing_logs)
			except Exception as e:
				# If JSON parsing fails, start fresh
				self.error_summary = json.dumps([log_entry])

		except Exception as e:
			frappe.log_error(f"Failed to add log: {str(e)}", "Transition Hub Import Session Log Error")

	def update_status(self, status, save=True):
		"""Update session status"""
		old_status = self.status
		self.status = status

		if status == "Completed":
			self.completed_at = now()
			self.progress_percentage = 100

		if save:
			try:
				self.save(ignore_permissions=True)
				frappe.db.commit()
			except Exception as e:
				frappe.log_error(f"Failed to save session: {str(e)}", "Session Save Error")

		# Publish real-time update
		try:
			frappe.publish_realtime(
				'import_session_update',
				{
					'session_id': self.name,
					'status': status,
					'old_status': old_status,
					'progress': self.progress_percentage,
					'success_count': self.success_count,
					'failed_count': self.failed_count,
					'duplicate_count': self.duplicate_count
				}
			)
		except Exception as e:
			frappe.log_error(f"Failed to publish update: {str(e)}", "Realtime Update Error")

	def mark_success(self, row_number=None, candidate_name=None, document_name=None,
					processing_time=None):
		"""Mark a row as successfully processed"""
		self.success_count = (self.success_count or 0) + 1
		self.add_simple_log(
			status="success",
			message=f"Successfully created candidate: {document_name}",
			row_number=row_number,
			candidate_name=candidate_name,
			document_name=document_name,
			processing_time=processing_time
		)

	def mark_failure(self, row_number=None, candidate_name=None, error_message=None,
					processing_time=None):
		"""Mark a row as failed"""
		self.failed_count = (self.failed_count or 0) + 1
		self.add_simple_log(
			status="error",
			message=error_message or "Processing failed",
			row_number=row_number,
			candidate_name=candidate_name,
			processing_time=processing_time
		)

	def mark_duplicate(self, row_number=None, candidate_name=None, message=None,
						processing_time=None):
		"""Mark a row as duplicate"""
		self.duplicate_count = (self.duplicate_count or 0) + 1
		self.add_simple_log(
			status="skipped",
			message=message or "Duplicate candidate skipped",
			row_number=row_number,
			candidate_name=candidate_name,
			processing_time=processing_time
		)

	def get_summary(self):
		"""Get import summary"""
		return {
			"session_id": self.name,
			"file_name": self.file_name,
			"status": self.status,
			"total_rows": self.total_rows,
			"success_count": self.success_count or 0,
			"failed_count": self.failed_count or 0,
			"duplicate_count": self.duplicate_count or 0,
			"progress_percentage": self.progress_percentage or 0,
			"started_at": self.started_at,
			"completed_at": self.completed_at,
			"estimated_time_remaining": self.estimated_time_remaining,
			"field_mapping": json.loads(self.field_mapping) if self.field_mapping else {}
		}

	def get_logs(self, limit=50):
		"""Get import logs"""
		logs = []
		if self.error_summary:
			try:
				all_logs = json.loads(self.error_summary)
				logs = all_logs[-limit:] if all_logs else []
			except:
				logs = []
		return logs

	def cancel_import(self, reason=None):
		"""Cancel the import session"""
		self.status = "Cancelled"
		self.completed_at = now()
		self.add_simple_log(
			status="info",
			message=f"Import cancelled. Reason: {reason or 'User cancelled'}"
		)
		self.save(ignore_permissions=True)

		# Publish cancellation
		frappe.publish_realtime(
			'import_cancelled',
			{
				'session_id': self.name,
				'reason': reason
			}
		)



# ===== EXCEL/CSV IMPORT FUNCTIONS =====

@frappe.whitelist()
def upload_and_preview():
    """
    Generic upload & preview API
    - Default doctype = Transition Hub Raw Candidate (backward compatible)
    - Supports CSV, XLS, XLSX
    """
    try:
        # =============================
        # 1. Resolve target doctype
        # =============================
        doctype = (
            frappe.form_dict.get("doctype")
            or frappe.request.form.get("doctype")
            or "Transition Hub Raw Candidate"   # backward compatibility
        )

        if not frappe.db.exists("DocType", doctype):
            frappe.throw(_("Invalid doctype: {0}").format(doctype))

        # =============================
        # 2. Get uploaded file
        # =============================
        uploaded = frappe.request.files.get("file")
        if not uploaded:
            frappe.throw(_("No file uploaded"))

        filename = uploaded.filename or ""
        ext = filename.split(".")[-1].lower()

        if ext not in ["csv", "xls", "xlsx"]:
            frappe.throw(
                _("Unsupported file type. Only CSV, XLS, XLSX files are allowed.")
            )

        # =============================
        # 3. Read file into DataFrame
        # =============================
        try:
            file_bytes = uploaded.read()
            bio = io.BytesIO(file_bytes)
            df = None

            if ext == "csv":
                encodings = ["utf-8", "latin-1", "cp1252", "iso-8859-1"]
                for encoding in encodings:
                    try:
                        bio.seek(0)
                        df = pd.read_csv(bio, encoding=encoding)
                        break
                    except (UnicodeDecodeError, UnicodeError):
                        continue

                if df is None:
                    frappe.throw(_("Unable to read CSV file. Please check file encoding."))

            else:  # xls / xlsx
                df = pd.read_excel(
                    bio,
                    engine="openpyxl" if ext == "xlsx" else None
                )

            if df is None or df.empty:
                frappe.throw(_("File is empty or cannot be read"))

            # Clean columns & rows
            df.columns = [str(col).strip() for col in df.columns]
            df = df.dropna(how="all")

            if df.empty:
                frappe.throw(_("No valid data found in file"))

        except Exception:
            frappe.log_error(frappe.get_traceback(), "Import Preview Read Error")
            frappe.throw(_("Failed to read file"))

        # =============================
        # 4. Prepare preview data
        # =============================
        columns = df.columns.tolist()
        sample = df.head(10).fillna("").to_dict(orient="records")

        # =============================
        # 5. Get available fields (GENERIC)
        # =============================
        meta = frappe.get_meta(doctype)
        available_fields = []

        for field in meta.fields:
            if field.fieldtype in (
                "Section Break",
                "Column Break",
                "Tab Break",
                "HTML",
            ):
                continue

            available_fields.append({
                "fieldname": field.fieldname,
                "label": field.label,
                "fieldtype": field.fieldtype,
                "reqd": field.reqd,
                "options": field.options,
            })

        # =============================
        # 6. Response (UNCHANGED STRUCTURE)
        # =============================
        return {
            "filename": filename,
            "doctype": doctype,
            "columns": columns,
            "sample": sample,
            "total_rows": len(df),
            "available_fields": available_fields,
        }

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Upload Preview Error")
        frappe.throw(_("Preview failed: {0}").format(str(e)))

def suggest_mapping(columns, meta):
    mapping = {}
    field_map = {
        f.fieldname.lower(): f.fieldname
        for f in meta.fields
    }

    for col in columns:
        key = col.lower().replace(" ", "_")
        if key in field_map:
            mapping[col] = field_map[key]

    return mapping


@frappe.whitelist()
def import_with_mapping():
    """
    Import with mapping
    - ATS_Candidate: giữ nguyên logic cũ
    - Transition Hub Raw Candidate: giữ nguyên logic cũ
    - Other Doctype: generic import + auto resolve Link field
    """
    try:
        uploaded = frappe.request.files.get('file')
        mapping_raw = frappe.form_dict.get('mapping', '{}')
        target_doctype = frappe.form_dict.get('doctype')
        batch_size = int(frappe.form_dict.get('batch_size', 100))
        validate_only = frappe.form_dict.get('validate_only', '0') == '1'
        skip_duplicates = frappe.form_dict.get('skip_duplicates', '1') == '1'
        job_opening = frappe.form_dict.get('job_opening')

        if not uploaded:
            frappe.throw(_("No file uploaded"))

        if not target_doctype:
            frappe.throw(_("Target Doctype is required"))

        mapping = json.loads(mapping_raw)
        if not mapping:
            frappe.throw(_("Field mapping is required"))

        # ===============================
        # Save file
        # ===============================
        file_doc = save_file(
            fname=uploaded.filename,
            content=uploaded.read(),
            dt="Transition Hub Import Session",
            dn=uploaded.filename,
            is_private=1
        )

        df = read_file_to_dataframe(
            file_doc.get_full_path(),
            uploaded.filename
        )

        # =========================================================
        # ATS CANDIDATE & TRANSITION HUB RAW CANDIDATE – GIỮ NGUYÊN LOGIC CŨ
        # =========================================================
        if target_doctype in ["ATS_Candidate", "Transition Hub Raw Candidate", "TransitionHubRawCandidate"]:

            session = create_import_session(
                file_name=uploaded.filename,
                file_url=file_doc.file_url,
                total_rows=len(df),
                import_type="Candidate" if target_doctype == "ATS_Candidate" else "Transition Hub Raw Candidate",
                target_doctype=target_doctype,
                job_opening=job_opening,
                field_mapping=mapping,
                batch_size=batch_size,
                validate_only=validate_only,
                skip_duplicates=skip_duplicates
            )

            session.add_simple_log(
                "info",
                f"{target_doctype} import started ({len(df)} rows)"
            )
            session.update_status("Processing")

            if validate_only:
                results = validate_import_data_with_session(
                    df, mapping, job_opening, session, target_doctype
                )
            else:
                results = process_import_data_with_session(
                    df, mapping, job_opening, session, batch_size, target_doctype
                )

            session.update_status("Completed")
            session.success_count = results.get("success", 0)
            session.failed_count = results.get("failed", 0)
            session.save(ignore_permissions=True)

            return {
                **results,
                "session_id": session.name
            }

        # =========================================================
        # GENERIC IMPORT – CHO MỌI DOCTYPE KHÁC
        # =========================================================

        def resolve_link_field(link_doctype, link_data, session):
            if not link_data:
                return None

            meta = frappe.get_meta(link_doctype)

            filters = {}
            for f in meta.fields:
                if f.reqd and f.fieldname in link_data:
                    filters[f.fieldname] = link_data[f.fieldname]

            if not filters:
                filters = link_data

            existing = frappe.db.get_value(link_doctype, filters)
            if existing:
                return existing

            doc = frappe.new_doc(link_doctype)
            for k, v in link_data.items():
                if hasattr(doc, k):
                    doc.set(k, v)

            doc.insert(ignore_permissions=True)

            session.add_simple_log(
                "info",
                f"Auto-created {link_doctype}: {doc.name}"
            )
            return doc.name

        session = create_import_session(
            file_name=uploaded.filename,
            file_url=file_doc.file_url,
            total_rows=len(df),
            import_type=target_doctype,
            target_doctype=target_doctype,
            field_mapping=mapping,
            batch_size=batch_size,
            validate_only=validate_only
        )

        session.update_status("Processing")

        meta = frappe.get_meta(target_doctype)
        success, failed = 0, 0
        row_results = []

        for idx, row in df.iterrows():
            try:
                doc = frappe.new_doc(target_doctype)

                for excel_col, map_conf in mapping.items():
                    value = row.get(excel_col)
                    if pd.isna(value) or value in (None, "", "nan", "NaN"):
                        continue

                    # Simple field mapping (string)
                    if isinstance(map_conf, str):
                        doc.set(map_conf, value)
                        continue

                    # Complex field mapping (dict with link fields)
                    if isinstance(map_conf, dict):
                        fieldname = map_conf.get("fieldname")
                        field = meta.get_field(fieldname)

                        if not field or field.fieldtype != "Link":
                            continue

                        link_doctype = field.options
                        link_data = {}

                        for k, excel_link_col in map_conf.get("link_fields", {}).items():
                            link_val = row.get(excel_link_col)
                            if not pd.isna(link_val) and link_val not in (None, "", "nan", "NaN"):
                                link_data[k] = link_val

                        link_name = resolve_link_field(
                            link_doctype,
                            link_data,
                            session
                        )

                        doc.set(fieldname, link_name)

                if validate_only:
                    doc.validate()
                else:
                    doc.insert(ignore_permissions=True)

                success += 1
                row_results.append({
                    "row": idx + 1,
                    "status": "success",
                    "name": doc.name
                })

            except Exception as e:
                failed += 1
                session.add_simple_log(
                    "error",
                    f"Row {idx+1}: {str(e)}"
                )
                row_results.append({
                    "row": idx + 1,
                    "status": "failed",
                    "error": str(e)
                })

        session.success_count = success
        session.failed_count = failed
        session.update_status("Completed")
        session.save(ignore_permissions=True)

        return {
            "session_id": session.name,
            "success": success,
            "failed": failed,
            "rows": row_results
        }

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Import With Mapping Error")
        frappe.throw(_("Import failed: {0}").format(str(e)))


def generic_validate_with_session(doctype, df, mapping, session):
    failed = 0

    for idx, row in df.iterrows():
        data = map_row_data(row, mapping)
        errors = []

        for field in mapping.values():
            if not data.get(field):
                errors.append(f"{field} is empty")

        if errors:
            failed += 1
            session.append("import_session_log", {
                "row_number": idx + 2,
                "status": "Error",
                "message": ", ".join(errors),
                "original_data": frappe.as_json(data)
            })

    session.save(ignore_permissions=True)
    return {
        "success": len(df) - failed,
        "failed": failed
    }

def generic_import_with_session(doctype, df, mapping, session, batch_size):
    success = failed = 0
    meta = frappe.get_meta(doctype)

    for idx, row in df.iterrows():
        data = map_row_data(row, mapping)

        try:
            doc = frappe.new_doc(doctype)
            doc.update(data)
            doc.insert(ignore_permissions=True)

            success += 1
            session.append("import_session_log", {
                "row_number": idx + 2,
                "status": "Success",
                "document_name": doc.name,
                "processed_data": frappe.as_json(doc.as_dict())
            })

        except Exception as e:
            failed += 1
            session.append("import_session_log", {
                "row_number": idx + 2,
                "status": "Error",
                "message": str(e),
                "error_details": frappe.get_traceback(),
                "original_data": frappe.as_json(data)
            })

        if (idx + 1) % batch_size == 0:
            session.save(ignore_permissions=True)

    session.save(ignore_permissions=True)
    return {
        "success": success,
        "failed": failed
    }


@frappe.whitelist()
def get_import_progress(session_id: str):
    """Get import progress for a session"""
    try:
        session = frappe.get_doc("Transition Hub Import Session", session_id)
        summary = session.get_summary()
        summary["logs"] = session.get_logs(50)  # Get last 50 logs
        return summary
    except frappe.DoesNotExistError:
        return {"status": "Not Found"}
    except Exception as e:
        frappe.log_error(f"Error getting import progress: {str(e)}", "Import Progress Error")
        return {"status": "Error", "message": str(e)}

@frappe.whitelist()
def download_import_template():
    """Generate and download import template Excel file"""
    try:
        # Get doctype from form data, default to Transition Hub Raw Candidate for backward compatibility
        target_doctype = (
            frappe.form_dict.get("doctype")
            or frappe.request.form.get("doctype")
            or "Transition Hub Raw Candidate"
        )

        if not frappe.db.exists("DocType", target_doctype):
            frappe.throw(_("Invalid doctype: {0}").format(target_doctype))

        # Get fields for the specified doctype
        meta = frappe.get_meta(target_doctype)

        # Get all non-system fields (avoid section breaks, column breaks, etc.)
        template_fields = []
        for field in meta.fields:
            if field.fieldtype not in [
                "Section Break", "Column Break", "Tab Break", "HTML",
                "Button", "Fold", "Heading"
            ] and not field.hidden and field.in_list_view:
                template_fields.append({
                    'fieldname': field.fieldname,
                    'label': field.label or field.fieldname,
                    'fieldtype': field.fieldtype,
                    'required': field.reqd,
                    'options': field.options
                })

        # Create template DataFrame with headers and sample data
        headers = []
        sample_data = {}

        for field in template_fields:
            header = f"{field['label']}"
            if field['required']:
                header += " *"
            headers.append(header)

            # Add sample data based on field type
            fieldname = field['fieldname']
            if field['fieldtype'] in ['Data', 'Text', 'Small Text', 'Long Text']:
                if fieldname == 'can_full_name':
                    sample_data[header] = 'Nguyen Van A'
                elif fieldname == 'can_email':
                    sample_data[header] = 'nguyenvana@example.com'
                elif fieldname == 'can_phone':
                    sample_data[header] = '+84901234567'
                elif fieldname == 'can_address':
                    sample_data[header] = '123 Phố Huế, Hai Bà Trưng, Hà Nội'
                else:
                    sample_data[header] = f'Sample {field["label"]}'
            elif field['fieldtype'] in ['Date']:
                sample_data[header] = '1990-01-01'
            elif field['fieldtype'] in ['Select']:
                options = field['options']
                if options:
                    option_list = options.split('\n')
                    # Skip the first empty element if present
                    option_list = [opt for opt in option_list if opt.strip()]
                    if option_list:
                        sample_data[header] = option_list[0]
                    else:
                        sample_data[header] = 'Option 1'
                else:
                    sample_data[header] = 'Select Value'
            elif field['fieldtype'] in ['Link']:
                sample_data[header] = f'{field["options"]} Name'
            elif field['fieldtype'] in ['Int', 'Float', 'Currency']:
                sample_data[header] = '123'
            elif field['fieldtype'] in ['Check']:
                sample_data[header] = '1'  # Checked
            else:
                sample_data[header] = f'Sample {field["label"]}'

        # Create DataFrame with sample data
        df = pd.DataFrame([sample_data]) if sample_data else pd.DataFrame()

        # Save to temporary file and return
        import tempfile

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            with pd.ExcelWriter(tmp.name, engine='openpyxl') as writer:
                sheet_name = f'{target_doctype.replace(" ", "_")} Template'
                df.to_excel(writer, sheet_name=sheet_name, index=False)

                # Get workbook and worksheet for styling
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]

                # Style headers
                from openpyxl.styles import Font, PatternFill, Alignment

                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                center_alignment = Alignment(horizontal="center", vertical="center")

                # Apply styling to header row
                for col_num, header in enumerate(headers, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment

                # Set column widths
                for col_num, header in enumerate(headers, 1):
                    column_letter = worksheet.cell(row=1, column=col_num).column_letter
                    worksheet.column_dimensions[column_letter].width = max(len(header) + 5, 15)

            # Read file content
            with open(tmp.name, 'rb') as f:
                content = f.read()

            # Clean up
            os.unlink(tmp.name)

        # Create file document
        timestamp = frappe.utils.now_datetime().strftime('%Y%m%d_%H%M%S')
        file_doc = frappe.get_doc({
            "doctype": "File",
            "file_name": f"{target_doctype.lower().replace(' ', '_')}_import_template_{timestamp}.xlsx",
            "content": content,
            "is_private": 0
        })
        file_doc.save(ignore_permissions=True)

        return {
            "file_url": file_doc.file_url,
            "file_name": file_doc.file_name
        }

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Template Generation Error")
        frappe.throw(_("Failed to generate template: {0}").format(str(e)))

# ===== HELPER FUNCTIONS =====

def create_import_session(file_name, file_url, total_rows, import_type="Candidate",
                         target_doctype=None, job_opening=None, field_mapping=None, batch_size=100,
                         validate_only=False, skip_duplicates=True, **kwargs):
    """Create a new import session"""
    session = frappe.new_doc("Transition Hub Import Session")
    session.file_name = file_name
    session.file_url = file_url
    session.total_rows = total_rows
    session.import_type = import_type
    session.target_doctype = target_doctype
    session.job_opening = job_opening
    session.field_mapping = json.dumps(field_mapping) if field_mapping else None
    session.status = "Pending"
    session.batch_size = batch_size
    session.validate_only = 1 if validate_only else 0
    session.skip_duplicates = 1 if skip_duplicates else 0
    session.created_by = frappe.session.user
    session.started_at = frappe.utils.now()

    # Set any additional kwargs
    for key, value in kwargs.items():
        if hasattr(session, key):
            setattr(session, key, value)

    session.insert(ignore_permissions=True)
    frappe.db.commit()
    return session

# ===== HELPER FUNCTIONS =====

def read_file_to_dataframe(file_path: str, filename: str) -> pd.DataFrame:
    """Read file into DataFrame with proper error handling"""
    ext = filename.split('.')[-1].lower()

    try:
        if ext in ['xlsx', 'xls']:
            df = pd.read_excel(file_path, engine='openpyxl' if ext == 'xlsx' else None)
        elif ext == 'csv':
            # Try different encodings
            encodings = ['utf-8', 'latin-1', 'cp1252']
            df = None
            for encoding in encodings:
                try:
                    df = pd.read_csv(file_path, encoding=encoding)
                    break
                except (UnicodeDecodeError, UnicodeError):
                    continue

            if df is None:
                frappe.throw(_("Cannot read CSV file with any supported encoding"))
        else:
            frappe.throw(_("Unsupported file format"))

        # Basic validation
        if df.empty:
            frappe.throw(_("File is empty"))

        # Clean data
        df = df.dropna(how='all')  # Remove empty rows
        df.columns = [str(col).strip() for col in df.columns]  # Clean column names

        return df

    except Exception as e:
        frappe.throw(_("Error reading file: {0}").format(str(e)))

def map_row_data(row: pd.Series, mapping: Dict) -> Dict:
    """Map row data according to field mapping"""
    candidate_data = {}

    for field_name, column_name in mapping.items():
        if column_name and column_name in row.index:
            value = row[column_name]
            if pd.notna(value) and str(value).strip():
                candidate_data[field_name] = str(value).strip()

    return candidate_data

def get_existing_emails(target_doctype: str = "ATS_Candidate"):
    """Get all existing candidate emails to prevent duplicates"""
    email_field = "can_email"  # Default email field name

    # For different doctypes, we might have different email field names
    if target_doctype == "TransitionHubRawCandidate":
        email_field = "email"  # Assuming TransitionHubRawCandidate uses 'email' field
    else:
        email_field = "can_email"  # Default for ATS_Candidate and others

    return set([
        email[0] for email in frappe.db.sql(
            f"SELECT {email_field} FROM `tab{target_doctype.replace(' ', '_')}` WHERE {email_field} IS NOT NULL AND {email_field} != ''"
        )
    ])

def validate_and_process_candidate(candidate_data: Dict, existing_emails: Set = None, target_doctype: str = "ATS_Candidate") -> Dict:
    """Enhanced validation with better error messages and field processing"""
    processed = {}
    errors = []

    # Determine required fields based on target doctype
    if target_doctype == "TransitionHubRawCandidate":
        # For TransitionHubRawCandidate, we might have different required fields
        # Let's make it flexible - check the doctype's required fields
        meta = frappe.get_meta(target_doctype)
        required_fields = {}
        for field in meta.fields:
            if field.reqd and field.fieldtype not in ["Table", "Table MultiSelect"]:
                required_fields[field.fieldname] = field.label or field.fieldname
    else:
        # Default for ATS_Candidate
        required_fields = {
            "can_full_name": _("Full Name"),
            "can_email": _("Email")
        }

    # Validate required fields
    for field, label in required_fields.items():
        value = candidate_data.get(field, "").strip()
        if not value:
            errors.append(_("Missing required field: {0}").format(label))

    # Email validation and duplicate check (only for candidate doctypes that have email)
    email = candidate_data.get("can_email", "")
    if email:
        email = email.strip().lower()
        if not validate_email(email):
            errors.append(_("Invalid email format: {0}").format(email))
        elif existing_emails and email in existing_emails:
            errors.append(_("Email already exists: {0}").format(email))
        else:
            processed["can_email"] = email
            if existing_emails is not None:
                existing_emails.add(email)

    # For TransitionHubRawCandidate, we might not need job opening
    job_id = candidate_data.get("job_opening_id") or candidate_data.get("can_email", "")
    if job_id and target_doctype != "TransitionHubRawCandidate":
        processed["job_opening_id"] = job_id

    # Determine field processors based on target doctype
    if target_doctype == "TransitionHubRawCandidate":
        # For raw candidate, we might have different fields - use generic processing
        # Or define specific processors for TransitionHubRawCandidate
        field_processors = {}
        # We'll process all fields generically for raw candidates
    else:
        # Default processors for ATS_Candidate
        field_processors = {
            "can_full_name": lambda x: process_text_field(x, 140),
            "can_phone": process_phone_number,
            "can_dob": process_date,
            "can_gender": process_gender,
            "can_address": lambda x: process_text_field(x, 500),
            "can_region": lambda x: process_text_field(x, 140),
            "candidatesource_id": lambda x: process_text_field(x, 140),
            "can_last_workplace": lambda x: process_text_field(x, 140),
            "can_other_links": lambda x: process_text_field(x, 500),
            "educationlevel_id": lambda x: process_text_field(x, 140),
            "institution_id": lambda x: process_text_field(x, 140),
            "major_id": lambda x: process_text_field(x, 140)
        }

    # Process fields based on target doctype
    if target_doctype == "TransitionHubRawCandidate":
        # For raw candidates, process all fields generically
        for field, value in candidate_data.items():
            if value and field not in ["job_opening_id", "can_application_date"]:  # Skip special fields
                processed[field] = str(value).strip() if value else None
    else:
        # For ATS_Candidate, use specific processors
        for field, processor in field_processors.items():
            if field in candidate_data and candidate_data[field]:
                try:
                    result = processor(candidate_data[field])
                    if result is not None:
                        processed[field] = result
                except Exception as e:
                    errors.append(_("Error processing {0}: {1}").format(field, str(e)))

    # Set default application date if needed
    if target_doctype != "TransitionHubRawCandidate":
        processed["can_application_date"] = candidate_data.get("can_application_date") or today()

    if errors:
        raise frappe.ValidationError("; ".join(errors))

    return processed

def validate_import_data_with_session(df: pd.DataFrame, mapping: Dict, job_opening: str = None, session=None, target_doctype: str = "ATS_Candidate") -> Dict:
    """Validate import data with session tracking"""
    results = {
        "success": 0,
        "failed": 0,
        "total": len(df),
        "logs": []
    }

    existing_emails = get_existing_emails(target_doctype)
    seen_emails = set()

    for idx, row in df.iterrows():
        start_time = time.time()
        row_number = idx + 1

        try:
            # Map fields
            candidate_data = map_row_data(row, mapping)
            if job_opening:
                candidate_data["job_opening_id"] = job_opening

            # Validate data
            processed_data = validate_and_process_candidate(candidate_data, existing_emails, target_doctype)

            # Check for duplicate email in current import
            email_key = "can_email" if target_doctype != "TransitionHubRawCandidate" else "email"
            email = processed_data.get(email_key, "").lower()
            if email and email in seen_emails:
                raise frappe.ValidationError(_("Duplicate email in import file: {0}").format(email))
            if email:
                seen_emails.add(email)

            processing_time = time.time() - start_time

            if session:
                session.mark_success(
                    row_number=row_number,
                    candidate_name=processed_data.get("can_full_name", processed_data.get("full_name", "Unknown")),
                    processing_time=processing_time
                )

            results["success"] += 1
            results["logs"].append({
                "row_number": row_number,
                "candidate_name": processed_data.get("can_full_name", processed_data.get("full_name", "Unknown")),
                "status": "success",
                "message": _("Validation passed")
            })

        except Exception as e:
            processing_time = time.time() - start_time
            error_msg = str(e)

            if session:
                session.mark_failure(
                    row_number=row_number,
                    candidate_name=candidate_data.get("can_full_name", candidate_data.get("full_name", "Unknown")) if 'candidate_data' in locals() else "Unknown",
                    error_message=error_msg,
                    processing_time=processing_time
                )

            results["failed"] += 1
            results["logs"].append({
                "row_number": row_number,
                "candidate_name": candidate_data.get("can_full_name", candidate_data.get("full_name", "Unknown")) if 'candidate_data' in locals() else "Unknown",
                "status": "error",
                "message": error_msg
            })

        # Update session progress periodically
        if session and idx % 10 == 0:
            session.save(ignore_permissions=True)
            frappe.db.commit()

    if session:
        session.save(ignore_permissions=True)
        frappe.db.commit()

    return results

def process_import_data_with_session(df: pd.DataFrame, mapping: Dict, job_opening: str, session, batch_size: int = 100, target_doctype: str = "ATS_Candidate") -> Dict:
    """Process and insert import data with session tracking"""
    results = {
        "success": 0,
        "failed": 0,
        "total": len(df),
        "logs": []
    }

    existing_emails = get_existing_emails(target_doctype)

    # Process in batches
    for start_idx in range(0, len(df), batch_size):
        end_idx = min(start_idx + batch_size, len(df))
        batch_df = df.iloc[start_idx:end_idx]

        batch_results = process_batch_with_session(
            batch_df, mapping, job_opening, existing_emails, start_idx, session, target_doctype
        )

        # Merge results
        results["success"] += batch_results["success"]
        results["failed"] += batch_results["failed"]
        results["logs"].extend(batch_results["logs"])

        # Update existing emails set for next batch
        existing_emails.update(batch_results.get("new_emails", []))

        # Commit batch and publish progress
        session.save(ignore_permissions=True)
        frappe.db.commit()

        # Publish real-time progress
        frappe.publish_realtime(
            'import_progress',
            {
                'session_id': session.name,
                'processed': end_idx,
                'total': len(df),
                'success': results["success"],
                'failed': results["failed"],
                'progress_percentage': (end_idx / len(df)) * 100
            }
        )

        # Small delay to prevent overwhelming the system
        time.sleep(0.1)

    return results

def process_batch_with_session(df: pd.DataFrame, mapping: Dict, job_opening: str,
                              existing_emails: Set, start_offset: int = 0, session=None, target_doctype: str = "ATS_Candidate") -> Dict:
    """Process a single batch with session tracking"""
    batch_results = {
        "success": 0,
        "failed": 0,
        "logs": [],
        "new_emails": []
    }

    for idx, row in df.iterrows():
        actual_row_num = start_offset + idx + 1
        start_time = time.time()

        try:
            # Map fields
            candidate_data = map_row_data(row, mapping)
            if job_opening:
                candidate_data["job_opening_id"] = job_opening


            # Validate and process
            processed_data = validate_and_process_candidate(candidate_data, existing_emails, target_doctype)

            # Create candidate document with the correct doctype
            doc = frappe.get_doc({
                "doctype": target_doctype,
                **processed_data
            })
            if job_opening and not hasattr(doc, 'job_opening_id') and 'job_opening_id' not in processed_data:
                processed_data["job_opening_id"] = job_opening
                doc.update({"job_opening_id": job_opening})

            doc.insert(ignore_permissions=True)

            processing_time = time.time() - start_time

            # Track new email
            email_key = "can_email" if target_doctype != "TransitionHubRawCandidate" else "email"
            email = processed_data.get(email_key)
            if email:
                batch_results["new_emails"].append(email.lower())

            if session:
                session.mark_success(
                    row_number=actual_row_num,
                    candidate_name=processed_data.get("can_full_name", processed_data.get("full_name", "Unknown")),
                    document_name=doc.name,
                    processing_time=processing_time
                )

            batch_results["success"] += 1
            batch_results["logs"].append({
                "row_number": actual_row_num,
                "candidate_name": processed_data.get("can_full_name", processed_data.get("full_name", "Unknown")),
                "status": "success",
                "message": _("Successfully created: {0}").format(doc.name),
                "document_name": doc.name
            })

        except Exception as e:
            processing_time = time.time() - start_time
            error_msg = str(e)

            if session:
                session.mark_failure(
                    row_number=actual_row_num,
                    candidate_name=candidate_data.get("can_full_name", candidate_data.get("full_name", "Unknown")) if 'candidate_data' in locals() else "Unknown",
                    error_message=error_msg,
                    processing_time=processing_time
                )

            batch_results["failed"] += 1
            batch_results["logs"].append({
                "row_number": actual_row_num,
                "candidate_name": candidate_data.get("can_full_name", candidate_data.get("full_name", "Unknown")) if 'candidate_data' in locals() else "Unknown",
                "status": "error",
                "message": error_msg
            })

    return batch_results

# ===== VALIDATION HELPER FUNCTIONS =====

def process_text_field(value: Any, max_length: int = None) -> str:
    """Process text field with length validation"""
    if not value:
        return None

    text = str(value).strip()
    if not text:
        return None

    if max_length and len(text) > max_length:
        text = text[:max_length]

    return text

def validate_email(email: str) -> bool:
    """Enhanced email validation"""
    if not email:
        return False

    # Basic format check
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    if not re.match(pattern, email):
        return False

    # Additional checks
    if len(email) > 254:  # Email length limit
        return False

    local, domain = email.split('@')
    if len(local) > 64:  # Local part length limit
        return False

    return True

def process_phone_number(phone: Any) -> Optional[str]:
    """Enhanced phone number processing"""
    if not phone:
        return None

    phone_str = str(phone).strip()
    if not phone_str:
        return None

    # Remove formatting but keep + for international numbers
    cleaned = re.sub(r'[^\d+]', '', phone_str)

    # Handle international format
    if cleaned.startswith('+'):
        cleaned = '+' + re.sub(r'[^\d]', '', cleaned[1:])
    else:
        cleaned = re.sub(r'[^\d]', '', cleaned)

    # Validate length
    digit_only = re.sub(r'[^\d]', '', cleaned)
    if len(digit_only) < 10 or len(digit_only) > 15:
        return None

    return cleaned[:20]  # Limit total length

def process_date(date_value: Any) -> Optional[str]:
    """Enhanced date processing with more formats"""
    if not date_value:
        return None

    if isinstance(date_value, datetime):
        return date_value.date().strftime('%Y-%m-%d')

    if isinstance(date_value, str):
        date_str = date_value.strip()
        if not date_str:
            return None

        # Extended date formats including Vietnamese
        date_formats = [
            "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y",
            "%Y/%m/%d", "%d.%m.%Y", "%Y.%m.%d", "%B %d, %Y",
            "%d/%m/%y", "%m/%d/%y", "%d-%m-%y", "%y-%m-%d"
        ]

        for fmt in date_formats:
            try:
                parsed_date = datetime.strptime(date_str, fmt).date()

                # Validation: reasonable birth date range
                current_year = datetime.now().year
                if parsed_date.year < 1940 or parsed_date.year > current_year:
                    continue

                if parsed_date > getdate():  # Not in future
                    continue

                return parsed_date.strftime('%Y-%m-%d')

            except ValueError:
                continue

    return None

def process_gender(gender: Any) -> Optional[str]:
    """Process gender with Vietnamese and English support"""
    if not gender:
        return None

    gender_str = str(gender).lower().strip()

    # Vietnamese mappings
    male_variants = ['nam', 'male', 'm', 'boy', 'man', '1']
    female_variants = ['nữ', 'nu', 'female', 'f', 'girl', 'woman', '0']

    if gender_str in male_variants:
        return 'Nam'
    elif gender_str in female_variants:
        return 'Nữ'

    return None

def validate_and_get_candidate_source(source):
    """Validate and return candidate source ID"""
    if frappe.db.exists("CandidateSource", source):
        return source
    return None

def validate_and_get_region(region):
    """Validate and return region ID"""
    if frappe.db.exists("Region", region):
        return region
    return None


@frappe.whitelist()
def retry_failed_import(session_id: str):
    """Retry failed rows from an import session"""
    try:
        session = frappe.get_doc("Transition Hub Import Session", session_id)

        if session.status not in ["Completed", "Failed"]:
            frappe.throw(_("Cannot retry an import that is still in progress"))

        # Get failed import logs
        failed_logs = frappe.get_all(
            "ImportLog",
            filters={
                "parent": session_id,
                "status": "Failed"
            },
            fields=["row_number", "candidate_name", "error_message"]
        )

        if not failed_logs:
            return {
                "retried": 0,
                "message": _("No failed rows to retry")
            }

        # Re-read the original file if it still exists
        original_file = frappe.db.get_value("File", {"file_url": session.file_url}, "name")
        if not original_file:
            frappe.throw(_("Original import file no longer exists. Cannot retry."))

        file_doc = frappe.get_doc("File", original_file)
        df = read_file_to_dataframe(file_doc.get_full_path(), session.file_name)

        # Get the field mapping
        mapping = json.loads(session.field_mapping) if session.field_mapping else {}
        if not mapping:
            frappe.throw(_("Field mapping not found in session"))

        # Create new retry session
        retry_session = create_import_session(
            file_name=f"RETRY_{session.file_name}",
            file_url=session.file_url,
            total_rows=len(failed_logs),
            import_type=session.import_type,
            target_doctype=session.target_doctype,
            job_opening=session.job_opening,
            field_mapping=mapping,
            batch_size=session.batch_size,
            validate_only=False,  # Always process for retry
            skip_duplicates=session.skip_duplicates,
            parent_session=session_id
        )

        retry_session.add_simple_log("info", f"Retry session started for {len(failed_logs)} failed rows from session {session_id}")
        retry_session.update_status("Processing")

        # Extract failed row numbers
        failed_row_numbers = [log.row_number - 1 for log in failed_logs]  # Convert to 0-based index
        failed_df = df.iloc[failed_row_numbers]

        # Process the failed rows
        results = process_import_data_with_session(
            failed_df,
            mapping,
            session.job_opening,
            retry_session,
            session.batch_size,
            session.target_doctype
        )

        # Update retry session status
        if results["failed"] == 0:
            retry_session.update_status("Completed")
            retry_session.add_simple_log("success", f"All {results['success']} rows processed successfully")
        else:
            retry_session.update_status("Completed")
            retry_session.error_summary = f"{results['failed']} rows still failed after retry"
            retry_session.add_simple_log("warning", f"{results['success']} succeeded, {results['failed']} still failed")

        return {
            "retried": len(failed_logs),
            "success": results["success"],
            "failed": results["failed"],
            "retry_session_id": retry_session.name,
            "message": _(f"Retry completed: {results['success']} successful, {results['failed']} still failed")
        }

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Import Retry Error")
        frappe.throw(_("Retry failed: {0}").format(str(e)))

@frappe.whitelist()
def get_retry_candidates(session_id: str):
    """Get list of failed rows that can be retried"""
    try:
        session = frappe.get_doc("Transition Hub Import Session", session_id)

        failed_logs = frappe.get_all(
            "ImportLog",
            filters={
                "parent": session_id,
                "status": "Failed"
            },
            fields=["row_number", "candidate_name", "error_message", "processing_time"],
            order_by="row_number"
        )

        return {
            "session_name": session.name,
            "total_failed": len(failed_logs),
            "failed_rows": failed_logs,
            "can_retry": len(failed_logs) > 0 and session.status in ["Completed", "Failed"]
        }

    except frappe.DoesNotExistError:
        return {"error": "Session not found"}
    except Exception as e:
        frappe.log_error(f"Error getting retry candidates: {str(e)}", "Retry Candidates Error")
        return {"error": str(e)}

@frappe.whitelist()
def cancel_import_session(session_id: str):
    """Cancel an active import session"""
    try:
        session = frappe.get_doc("Transition Hub Import Session", session_id)

        if session.status not in ["Processing", "Pending"]:
            frappe.throw(_("Can only cancel sessions that are pending or in progress"))

        session.update_status("Cancelled")
        session.add_simple_log("warning", "Import session cancelled by user")

        # Note: This won't stop background jobs that are already running
        # You might need additional logic to handle job cancellation

        return {
            "message": _("Import session cancelled successfully"),
            "status": "Cancelled"
        }

    except Exception as e:
        frappe.log_error(f"Error cancelling import session: {str(e)}", "Cancel Import Error")
        frappe.throw(_("Failed to cancel import: {0}").format(str(e)))

@frappe.whitelist()
def get_import_sessions():
    """Get list of import sessions with summary"""
    try:
        sessions = frappe.get_all(
            "Transition Hub Import Session",
            fields=[
                "name", "file_name", "created_by", "started_at", "completed_at",
                "status", "total_rows", "success_count", "failed_count",
                "import_type", "target_doctype", "job_opening", "validate_only", "error_summary"
            ],
            order_by="started_at desc",
            limit=50
        )

        # Add retry information
        for session in sessions:
            # Check if this is a retry session
            if session.get("file_name", "").startswith("RETRY_"):
                session["is_retry"] = True
                # Get parent session if available
                parent_session = frappe.db.get_value("Transition Hub Import Session",
                    {"name": session.name}, "parent_session")
                if parent_session:
                    session["parent_session"] = parent_session
            else:
                session["is_retry"] = False

            # Count failed rows for retry possibility
            if session.status in ["Completed", "Failed"]:
                failed_count = frappe.db.count("ImportLog", {
                    "parent": session.name,
                    "status": "Failed"
                })
                session["can_retry"] = failed_count > 0
                session["retry_candidates"] = failed_count
            else:
                session["can_retry"] = False
                session["retry_candidates"] = 0

        return sessions

    except Exception as e:
        frappe.log_error(f"Error getting import sessions: {str(e)}", "Get Import Sessions Error")
        return []

@frappe.whitelist()
def delete_import_session(session_id: str):
    """Delete an import session and its logs"""
    try:
        session = frappe.get_doc("Transition Hub Import Session", session_id)

        if session.status == "Processing":
            frappe.throw(_("Cannot delete a session that is currently processing"))

        # Delete associated logs first
        frappe.db.delete("ImportLog", {"parent": session_id})

        # Delete the session
        frappe.delete_doc("Transition Hub Import Session", session_id, ignore_permissions=True)
        frappe.db.commit()

        return {
            "message": _("Import session deleted successfully")
        }

    except Exception as e:
        frappe.log_error(f"Error deleting import session: {str(e)}", "Delete Import Session Error")
        frappe.throw(_("Failed to delete session: {0}").format(str(e)))

@frappe.whitelist()
def export_failed_rows(session_id: str):
    """Export failed rows to Excel for manual correction"""
    try:
        session = frappe.get_doc("Transition Hub Import Session", session_id)

        # Get failed logs with details
        failed_logs = frappe.get_all(
            "ImportLog",
            filters={
                "parent": session_id,
                "status": "Failed"
            },
            fields=["row_number", "candidate_name", "error_message", "raw_data"],
            order_by="row_number"
        )

        if not failed_logs:
            frappe.throw(_("No failed rows found to export"))

        # Try to reconstruct the original data
        original_file = frappe.db.get_value("File", {"file_url": session.file_url}, "name")
        if original_file:
            try:
                file_doc = frappe.get_doc("File", original_file)
                df = read_file_to_dataframe(file_doc.get_full_path(), session.file_name)

                # Extract failed rows
                failed_row_numbers = [log.row_number - 1 for log in failed_logs]
                failed_df = df.iloc[failed_row_numbers].copy()

                # Add error information
                error_messages = [log.error_message for log in failed_logs]
                failed_df.insert(0, 'Row_Number', [log.row_number for log in failed_logs])
                failed_df.insert(1, 'Error_Message', error_messages)

                # Create Excel file
                import tempfile
                with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                    with pd.ExcelWriter(tmp.name, engine='openpyxl') as writer:
                        failed_df.to_excel(writer, sheet_name='Failed Rows', index=False)

                        # Style the worksheet
                        from openpyxl.styles import Font, PatternFill
                        workbook = writer.book
                        worksheet = writer.sheets['Failed Rows']

                        # Style error columns
                        error_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                        header_font = Font(bold=True)

                        # Apply styling
                        for col in range(1, 3):  # Row_Number and Error_Message columns
                            for row in range(2, len(failed_df) + 2):
                                worksheet.cell(row=row, column=col).fill = error_fill

                        # Style headers
                        for col in range(1, len(failed_df.columns) + 1):
                            worksheet.cell(row=1, column=col).font = header_font

                    # Read the file
                    with open(tmp.name, 'rb') as f:
                        content = f.read()

                    # Clean up
                    os.unlink(tmp.name)

                # Create Frappe file
                file_doc = frappe.get_doc({
                    "doctype": "File",
                    "file_name": f"failed_rows_{session_id}_{frappe.utils.now_datetime().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    "content": content,
                    "is_private": 0
                })
                file_doc.save(ignore_permissions=True)

                return {
                    "file_url": file_doc.file_url,
                    "file_name": file_doc.file_name,
                    "failed_count": len(failed_logs)
                }

            except Exception as e:
                frappe.log_error(f"Error creating failed rows export: {str(e)}")
                # Fallback to simple error list
                pass

        # Fallback: Create simple error summary
        error_data = []
        for log in failed_logs:
            error_data.append({
                'Row_Number': log.row_number,
                'Candidate_Name': log.candidate_name or 'Unknown',
                'Error_Message': log.error_message
            })

        df = pd.DataFrame(error_data)

        # Create Excel file
        import tempfile
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            df.to_excel(tmp.name, index=False)

            with open(tmp.name, 'rb') as f:
                content = f.read()

            os.unlink(tmp.name)

        # Create Frappe file
        file_doc = frappe.get_doc({
            "doctype": "File",
            "file_name": f"import_errors_{session_id}_{frappe.utils.now_datetime().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "content": content,
            "is_private": 0
        })
        file_doc.save(ignore_permissions=True)

        return {
            "file_url": file_doc.file_url,
            "file_name": file_doc.file_name,
            "failed_count": len(failed_logs)
        }

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Export Failed Rows Error")
        frappe.throw(_("Failed to export failed rows: {0}").format(str(e)))
